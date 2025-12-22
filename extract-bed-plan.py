import openpyxl
import json
from datetime import datetime, date

# Custom JSON encoder for datetime objects
class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)

def fmt_date(val):
    """Format a date value to ISO string or None"""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if hasattr(val, 'isoformat'):
        return val.isoformat()
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%dT%H:%M:%S')
    return str(val) if val else None

def fmt_number(val):
    """Format a numeric value, handling Excel date encoding quirk"""
    if val is None:
        return None
    # Excel sometimes encodes small numbers as dates starting from 1900-01-01
    # If we get a datetime in year 1900, convert it back to a number (days since epoch)
    if isinstance(val, (datetime, date)):
        if val.year == 1900:
            # Excel epoch is 1900-01-01, but has a leap year bug, so day 1 = Jan 1 1900
            # The number stored is (date - 1900-01-01).days
            epoch = date(1900, 1, 1)
            if isinstance(val, datetime):
                val = val.date()
            days = (val - epoch).days
            return days
        return None  # Actual dates shouldn't be numbers
    try:
        return float(val) if val else None
    except (ValueError, TypeError):
        return None

# Load the workbook
wb = openpyxl.load_workbook("Crop Plan 2025 V20.xlsm", data_only=True)

# Read the Bed Plan sheet
ws = wb["Bed Plan"]

# Extract bed assignments with timing data
bed_assignments = []

for row in range(6, ws.max_row + 1):  # Data starts at row 6
    crop_name = ws.cell(row=row, column=1).value  # Column A
    identifier = ws.cell(row=row, column=2).value  # Column B
    bed = ws.cell(row=row, column=3).value  # Column C

    if crop_name and identifier and bed:
        assignment = {
            "crop": crop_name,
            "identifier": identifier,
            "bed": bed,
            # Basic info
            "bedsCount": ws.cell(row=row, column=14).value,  # # Of Beds

            # Timing - Start dates
            "startDate": fmt_date(ws.cell(row=row, column=16).value),  # Start Date (the seed/planning anchor)
            "plannedGreenhouseStartDate": fmt_date(ws.cell(row=row, column=17).value),
            "actualGreenhouseDate": fmt_date(ws.cell(row=row, column=18).value),
            "greenhouseStartDate": fmt_date(ws.cell(row=row, column=19).value),  # Computed GH start
            "fixedFieldStartDate": fmt_date(ws.cell(row=row, column=20).value),  # Manual override for field date
            "followsCrop": ws.cell(row=row, column=21).value,
            "followOffset": ws.cell(row=row, column=22).value,

            # Timing - Field dates
            "plannedTpOrDsDate": fmt_date(ws.cell(row=row, column=23).value),  # Default transplant/direct seed date
            "actualTpOrDsDate": fmt_date(ws.cell(row=row, column=24).value),
            "inGroundDaysLate": ws.cell(row=row, column=25).value,  # Deviation from plan
            "tpOrDsDate": fmt_date(ws.cell(row=row, column=26).value),  # Final computed TP/DS date
            "additionalDaysInField": ws.cell(row=row, column=27).value,  # ADDITIONAL DAYS

            # Timing - Harvest dates
            "expectedBeginningOfHarvest": fmt_date(ws.cell(row=row, column=28).value),
            "actualBeginningOfHarvest": fmt_date(ws.cell(row=row, column=29).value),
            "beginningOfHarvest": fmt_date(ws.cell(row=row, column=30).value),  # Final computed
            "expectedEndOfHarvest": fmt_date(ws.cell(row=row, column=31).value),  # From normalized calc
            "actualEndOfHarvest": fmt_date(ws.cell(row=row, column=32).value),
            "additionalDaysOfHarvest": ws.cell(row=row, column=33).value,  # ADDITIONAL DAYS
            "augustHarvest": ws.cell(row=row, column=34).value,
            "failed": ws.cell(row=row, column=35).value,
            "endOfHarvest": fmt_date(ws.cell(row=row, column=36).value),  # Final computed end

            # Config values from crop database
            "dtm": ws.cell(row=row, column=57).value,
            "additionalDaysInCells": ws.cell(row=row, column=58).value,  # ADDITIONAL DAYS
            "harvestWindow": ws.cell(row=row, column=59).value,
            "dsTp": ws.cell(row=row, column=60).value,  # Direct Seed or Transplant
            "daysInCells": ws.cell(row=row, column=62).value,  # Base days in cells
            "daysUntilHarvest": ws.cell(row=row, column=71).value,
            "trueHarvestWindow": ws.cell(row=row, column=72).value,
            "category": ws.cell(row=row, column=73).value,
            "growingStructure": ws.cell(row=row, column=87).value,
        }
        bed_assignments.append(assignment)

print(f"Found {len(bed_assignments)} bed assignments")

# Show unique beds
beds = sorted(set(a["bed"] for a in bed_assignments))
print(f"\nUnique beds ({len(beds)}): {beds}")

# Group beds by letter prefix (row/section)
bed_groups = {}
for bed in beds:
    # Extract letter prefix (e.g., "A" from "A5", "GH" from "GH1")
    prefix = ""
    for char in bed:
        if char.isalpha():
            prefix += char
        else:
            break
    if prefix not in bed_groups:
        bed_groups[prefix] = []
    bed_groups[prefix].append(bed)

print(f"\nBed groups:")
for prefix, group_beds in sorted(bed_groups.items()):
    print(f"  {prefix}: {group_beds}")

# Save to JSON
with open("crop-api/src/data/bed-plan.json", "w") as f:
    json.dump({
        "assignments": bed_assignments,
        "beds": beds,
        "bedGroups": {k: v for k, v in sorted(bed_groups.items())}
    }, f, indent=2, cls=DateTimeEncoder)

print(f"\nSaved to crop-api/src/data/bed-plan.json")

# Show sample assignments with timing
print("\nSample assignments with ALL timing adjustments:")
for a in bed_assignments[:5]:
    print(f"\n  {a['identifier']} -> {a['bed']}: {a['crop'][:40]}")
    print(f"    Start: {a['startDate']}")
    print(f"    TP/DS: {a['tpOrDsDate']} (planned: {a['plannedTpOrDsDate']})")
    print(f"    Harvest: {a['beginningOfHarvest']} - {a['endOfHarvest']}")
    print(f"    Expected End: {a['expectedEndOfHarvest']}")
    print(f"    ADJUSTMENTS:")
    print(f"      Additional Days in Cells: {a['additionalDaysInCells']}")
    print(f"      Additional Days In Field: {a['additionalDaysInField']}")
    print(f"      Additional Days of Harvest: {a['additionalDaysOfHarvest']}")
    print(f"      In Ground Days Late: {a['inGroundDaysLate']}")

# Summary of adjustments
print("\n\n=== ADJUSTMENT SUMMARY ===")
crops_with_cell_adj = [a for a in bed_assignments if a['additionalDaysInCells']]
crops_with_field_adj = [a for a in bed_assignments if a['additionalDaysInField']]
crops_with_harvest_adj = [a for a in bed_assignments if a['additionalDaysOfHarvest']]
crops_late = [a for a in bed_assignments if a['inGroundDaysLate']]

print(f"Crops with Additional Days in Cells: {len(crops_with_cell_adj)}")
print(f"Crops with Additional Days in Field: {len(crops_with_field_adj)}")
print(f"Crops with Additional Days of Harvest: {len(crops_with_harvest_adj)}")
print(f"Crops with In Ground Days Late: {len(crops_late)}")

if crops_with_harvest_adj:
    print(f"\nCrops with harvest extensions:")
    for a in crops_with_harvest_adj[:10]:
        print(f"  {a['identifier']}: +{a['additionalDaysOfHarvest']} days harvest")
