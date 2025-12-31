#!/usr/bin/env python3
"""
Trace Excel formula dependencies and build a DAG.

Analyzes Excel formulas that use structured table references
(TableName[[#This Row],[Column]]) and outputs dependency graphs.

Features:
- Scans ALL rows to detect formula variance per column
- Detects cross-table references (XLOOKUP, INDEX/MATCH)
- Reports base formula vs deviations
- Outputs qualified dependencies (e.g., CropChart.DTM)

Usage:
    python trace-formula-dag.py <workbook> <sheet> [options]

Examples:
    python trace-formula-dag.py "Workbook.xlsx" "Data" --columns 16-36
    python trace-formula-dag.py "Plan.xlsx" "Sheet1" --table MyTable --columns 1-20,45,46
    python trace-formula-dag.py "Data.xlsx" "Main" --header-row 3 --all
"""

import argparse
import openpyxl
import re
import sys
import json
from collections import defaultdict, Counter


def parse_column_spec(spec):
    """
    Parse a column specification string into a list of column numbers.

    Supports:
    - Single columns: "5"
    - Ranges: "16-36"
    - Comma-separated: "16-36,57,58,59"
    - "all" for all columns with formulas
    """
    if spec.lower() == "all":
        return None  # Signal to analyze all columns

    columns = []
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-")
            columns.extend(range(int(start), int(end) + 1))
        else:
            columns.append(int(part))
    return columns


def normalize_formula(formula):
    """
    Normalize a formula for comparison purposes.
    Strips whitespace variations but preserves structure.
    """
    if not formula:
        return None
    return re.sub(r'\s+', ' ', str(formula).strip())


def parse_structured_refs(formula, table_name):
    """
    Extract column references from structured table references.
    Handles both same-row and full-column reference patterns.
    Returns list of (table_name, column_name) tuples.
    """
    if not formula or not str(formula).startswith("="):
        return []

    refs = []
    formula_str = str(formula)

    # Same-row references: Table[[#This Row],[Column Name]]
    same_row_pattern = rf"({re.escape(table_name)})\[\[#This Row\],\[([^\]]+)\]\]"
    for match in re.finditer(same_row_pattern, formula_str):
        refs.append((match.group(1), match.group(2)))

    # Short form same-row: [@[Column Name]] (implicit current table)
    short_row_pattern = r"@\[([^\]]+)\]"
    for match in re.finditer(short_row_pattern, formula_str):
        refs.append((table_name, match.group(1)))

    # Full column references: Table[Column Name] (not preceded by @)
    col_pattern = rf"(?<!@)({re.escape(table_name)})\[([^\[\]@]+)\]"
    for match in re.finditer(col_pattern, formula_str):
        refs.append((match.group(1), match.group(2)))

    return list(set(refs))


def parse_cross_table_refs(formula, known_tables):
    """
    Extract cross-table references from XLOOKUP, INDEX/MATCH patterns.
    Returns list of (table_name, column_name, ref_type) tuples.
    """
    if not formula or not str(formula).startswith("="):
        return []

    refs = []
    formula_str = str(formula)

    for table in known_tables:
        # XLOOKUP pattern: XLOOKUP(key, Table[KeyCol], Table[ValueCol])
        # We want to capture references to other tables
        xlookup_pattern = rf"XLOOKUP\s*\([^,]+,\s*{re.escape(table)}\[([^\]]+)\]\s*,\s*{re.escape(table)}\[([^\]]+)\]"
        for match in re.finditer(xlookup_pattern, formula_str, re.IGNORECASE):
            refs.append((table, match.group(1), "xlookup_key"))
            refs.append((table, match.group(2), "xlookup_value"))

        # INDEX/MATCH pattern: INDEX(Table[Col], MATCH(...))
        index_pattern = rf"INDEX\s*\(\s*{re.escape(table)}\[([^\]]+)\]"
        for match in re.finditer(index_pattern, formula_str, re.IGNORECASE):
            refs.append((table, match.group(1), "index"))

        # General table column reference (for other tables)
        general_pattern = rf"(?<![a-zA-Z]){re.escape(table)}\[([^\[\]]+)\]"
        for match in re.finditer(general_pattern, formula_str):
            refs.append((table, match.group(1), "reference"))

    return list(set(refs))


def build_header_map(ws, header_row):
    """Build bidirectional header mapping."""
    headers = {}
    col_to_header = {}

    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[val] = col
            col_to_header[col] = val

    return headers, col_to_header


def find_data_range(ws, header_row):
    """Find the first and last data rows (non-empty rows after header)."""
    first_data_row = header_row + 1
    last_data_row = first_data_row

    # Find last row with data in column A (or first non-empty column)
    for row in range(first_data_row, ws.max_row + 1):
        has_data = False
        for col in range(1, min(10, ws.max_column + 1)):  # Check first 10 cols
            if ws.cell(row=row, column=col).value is not None:
                has_data = True
                break
        if has_data:
            last_data_row = row
        elif row > last_data_row + 5:  # Allow small gaps
            break

    return first_data_row, last_data_row


def analyze_column_formulas(ws, col_num, header_row, first_row, last_row):
    """
    Analyze all formulas in a column and return variance stats.

    Returns:
        dict with:
        - total_rows: number of data rows
        - formula_rows: rows with formulas
        - value_rows: rows with static values
        - empty_rows: rows with no value
        - formulas: Counter of normalized formulas
        - base_formula: most common formula (or None)
        - variance_pct: percentage of rows that deviate from base
        - sample_values: sample of static values if mostly values
    """
    formulas = Counter()
    value_count = 0
    empty_count = 0
    sample_values = []

    for row in range(first_row, last_row + 1):
        cell = ws.cell(row=row, column=col_num)
        val = cell.value

        if val is None or val == "":
            empty_count += 1
        elif str(val).startswith("="):
            normalized = normalize_formula(val)
            formulas[normalized] += 1
        else:
            value_count += 1
            if len(sample_values) < 5:
                sample_values.append(val)

    total_rows = last_row - first_row + 1
    formula_rows = sum(formulas.values())

    # Determine base formula (most common)
    base_formula = None
    base_count = 0
    if formulas:
        base_formula, base_count = formulas.most_common(1)[0]

    # Calculate variance
    variance_pct = 0
    if formula_rows > 0:
        variance_pct = round((formula_rows - base_count) / formula_rows * 100, 1)

    return {
        "total_rows": total_rows,
        "formula_rows": formula_rows,
        "value_rows": value_count,
        "empty_rows": empty_count,
        "formulas": formulas,
        "base_formula": base_formula,
        "base_count": base_count,
        "variance_pct": variance_pct,
        "sample_values": sample_values,
        "unique_formulas": len(formulas)
    }


def classify_column(analysis):
    """
    Classify a column based on its formula analysis.

    Returns one of:
    - INPUT: No formulas, all manual values
    - CALCULATED: Has formulas, consistent across rows
    - MIXED: Has formulas but significant variance
    - EMPTY: Mostly empty
    """
    total = analysis["total_rows"]
    formulas = analysis["formula_rows"]
    values = analysis["value_rows"]
    empty = analysis["empty_rows"]
    variance = analysis["variance_pct"]

    if empty > total * 0.9:
        return "EMPTY"
    if formulas == 0:
        return "INPUT"
    if variance < 5:  # Less than 5% deviation
        return "CALCULATED"
    return "MIXED"


def get_all_dependencies(analysis, table_name, other_tables):
    """
    Extract all dependencies from all formula variants in a column.
    Returns dict with 'internal' (same table) and 'external' (other tables) deps.
    """
    internal_deps = set()
    external_deps = set()

    for formula in analysis["formulas"].keys():
        if not formula:
            continue

        # Same-table references
        refs = parse_structured_refs(formula, table_name)
        for tbl, col in refs:
            if tbl == table_name:
                internal_deps.add(col)

        # Cross-table references
        cross_refs = parse_cross_table_refs(formula, other_tables)
        for tbl, col, ref_type in cross_refs:
            if tbl != table_name:
                external_deps.add(f"{tbl}.{col}")

    return {
        "internal": sorted(internal_deps),
        "external": sorted(external_deps)
    }


def build_dag(ws, headers, col_to_header, target_columns, table_name,
              other_tables, header_row, first_row, last_row):
    """Build the complete dependency DAG for specified columns."""
    dag = {}

    for col in target_columns:
        header = col_to_header.get(col, f"Col {col}")

        # Analyze all rows for this column
        analysis = analyze_column_formulas(ws, col, header_row, first_row, last_row)
        classification = classify_column(analysis)
        deps = get_all_dependencies(analysis, table_name, other_tables)

        # Map internal dep names to column numbers
        internal_col_nums = [headers[d] for d in deps["internal"] if d in headers]

        dag[col] = {
            "header": header,
            "classification": classification,
            "base_formula": analysis["base_formula"],
            "formula_rows": analysis["formula_rows"],
            "value_rows": analysis["value_rows"],
            "variance_pct": analysis["variance_pct"],
            "unique_formulas": analysis["unique_formulas"],
            "depends_on": internal_col_nums,
            "external_deps": deps["external"],
            "sample_values": analysis["sample_values"]
        }

    return dag


def get_level(col, dag, cache=None):
    """Calculate the dependency level of a column (0 = input)."""
    if cache is None:
        cache = {}
    if col in cache:
        return cache[col]

    info = dag.get(col, {})
    deps = info.get("depends_on", [])
    external = info.get("external_deps", [])

    # If no internal deps but has external deps, it's level 1
    if not deps:
        if external:
            cache[col] = 1
        else:
            cache[col] = 0
        return cache[col]

    valid_deps = [d for d in deps if d in dag]
    if not valid_deps:
        cache[col] = 1
        return 1

    max_dep_level = max(get_level(d, dag, cache) for d in valid_deps)
    cache[col] = max_dep_level + 1
    return cache[col]


def print_tree(col, dag, col_to_header, indent=0, visited=None):
    """Print dependency tree recursively."""
    if visited is None:
        visited = set()
    if col in visited:
        print("  " * indent + f"|- [{col}] (circular ref)")
        return
    visited.add(col)

    info = dag.get(col, {})
    header = info.get("header", f"Col {col}")
    classification = info.get("classification", "UNKNOWN")
    formula = info.get("base_formula")
    deps = info.get("depends_on", [])
    external = info.get("external_deps", [])
    variance = info.get("variance_pct", 0)

    prefix = "  " * indent + ("|- " if indent > 0 else "")

    # Status indicator
    status = ""
    if variance > 0:
        status = f" ⚠️ {variance}% variance"

    if classification == "INPUT":
        print(f"{prefix}[{col}] {header} (INPUT)")
    elif classification == "EMPTY":
        print(f"{prefix}[{col}] {header} (EMPTY)")
    else:
        formula_short = formula[:60] + "..." if formula and len(formula) > 60 else formula
        print(f"{prefix}[{col}] {header}{status}")
        if formula_short:
            print(f"{prefix}     = {formula_short}")

    # Show external dependencies
    if external:
        print(f"{prefix}     → External: {', '.join(external)}")

    # Recurse into internal dependencies
    for dep in deps:
        print_tree(dep, dag, col_to_header, indent + 1, visited.copy())


def find_columns_with_content(ws, header_row, first_row, last_row):
    """Find all columns that have any content (formulas or values)."""
    content_cols = []
    for col in range(1, ws.max_column + 1):
        for row in range(first_row, min(first_row + 10, last_row + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                content_cols.append(col)
                break
    return content_cols


def main():
    parser = argparse.ArgumentParser(
        description="Trace Excel formula dependencies and build a DAG.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s "Workbook.xlsx" "Sheet1" --columns 16-36
  %(prog)s "Plan.xlsx" "Data" --table MyTable --columns 1-20,45,46
  %(prog)s "Data.xlsx" "Main" --all --header-row 3
  %(prog)s "Plan.xlsx" "BedPlan" --table BedPlan --other-tables CropChart
        """
    )

    parser.add_argument("workbook", help="Path to the Excel workbook")
    parser.add_argument("sheet", help="Name of the sheet to analyze")
    parser.add_argument(
        "--table", "-t",
        default="Table1",
        help="Name of the Excel table (default: Table1)"
    )
    parser.add_argument(
        "--other-tables", "-o",
        default="",
        help="Comma-separated list of other table names to detect cross-references"
    )
    parser.add_argument(
        "--columns", "-c",
        default="all",
        help="Columns to analyze: ranges (16-36), lists (1,5,10), or 'all' (default: all)"
    )
    parser.add_argument(
        "--header-row", "-r",
        type=int,
        default=1,
        help="Row number containing headers (default: 1)"
    )
    parser.add_argument(
        "--target", "-T",
        type=int,
        help="Specific column to trace from (prints tree starting from this column)"
    )
    parser.add_argument(
        "--quiet", "-q",
        action="store_true",
        help="Only print the dependency tree, not full DAG"
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output DAG as JSON for programmatic use"
    )
    parser.add_argument(
        "--inputs-only",
        action="store_true",
        help="Only show INPUT columns (entry points)"
    )

    args = parser.parse_args()

    other_tables = [t.strip() for t in args.other_tables.split(",") if t.strip()]

    # Load workbook with formulas (not values)
    print(f"Loading: {args.workbook}", file=sys.stderr)
    print(f"Sheet: {args.sheet}, Table: {args.table}", file=sys.stderr)
    if other_tables:
        print(f"Cross-table detection: {other_tables}", file=sys.stderr)

    try:
        wb = openpyxl.load_workbook(args.workbook, data_only=False)
        ws = wb[args.sheet]
    except Exception as e:
        print(f"Error loading workbook: {e}", file=sys.stderr)
        sys.exit(1)

    # Build header mappings
    headers, col_to_header = build_header_map(ws, args.header_row)
    print(f"Found {len(headers)} headers", file=sys.stderr)

    # Find data range
    first_row, last_row = find_data_range(ws, args.header_row)
    print(f"Data rows: {first_row} to {last_row} ({last_row - first_row + 1} rows)", file=sys.stderr)

    # Determine columns to analyze
    target_columns = parse_column_spec(args.columns)
    if target_columns is None:
        target_columns = find_columns_with_content(ws, args.header_row, first_row, last_row)

    print(f"Analyzing {len(target_columns)} columns...", file=sys.stderr)

    # Build the DAG
    dag = build_dag(ws, headers, col_to_header, target_columns, args.table,
                    other_tables, args.header_row, first_row, last_row)

    # JSON output mode
    if args.json:
        output = {
            "table": args.table,
            "sheet": args.sheet,
            "data_rows": last_row - first_row + 1,
            "columns": {}
        }
        for col, info in dag.items():
            output["columns"][col] = {
                "header": info["header"],
                "classification": info["classification"],
                "base_formula": info["base_formula"],
                "formula_rows": info["formula_rows"],
                "value_rows": info["value_rows"],
                "unique_formulas": info["unique_formulas"],
                "variance_pct": info["variance_pct"],
                "depends_on": [col_to_header.get(d, f"Col {d}") for d in info["depends_on"]],
                "external_deps": info["external_deps"]
            }
        print(json.dumps(output, indent=2))
        return

    # Inputs-only mode
    if args.inputs_only:
        print("\n" + "=" * 80)
        print("INPUT COLUMNS (Entry Points)")
        print("=" * 80)

        for col in sorted(dag.keys()):
            info = dag[col]
            if info["classification"] == "INPUT":
                header = info["header"]
                samples = info.get("sample_values", [])
                sample_str = f" (e.g., {samples[0]})" if samples else ""
                print(f"  [{col:2d}] {header}{sample_str}")
        return

    if not args.quiet:
        # Print summary by classification
        print("\n" + "=" * 80)
        print("COLUMN CLASSIFICATION SUMMARY")
        print("=" * 80)

        by_class = defaultdict(list)
        for col, info in dag.items():
            by_class[info["classification"]].append((col, info["header"]))

        for cls in ["INPUT", "CALCULATED", "MIXED", "EMPTY"]:
            if cls in by_class:
                print(f"\n{cls} ({len(by_class[cls])} columns):")
                for col, header in sorted(by_class[cls]):
                    info = dag[col]
                    extras = []
                    if info["variance_pct"] > 0:
                        extras.append(f"{info['variance_pct']}% variance")
                    if info["unique_formulas"] > 1:
                        extras.append(f"{info['unique_formulas']} variants")
                    if info["external_deps"]:
                        extras.append(f"→ {', '.join(info['external_deps'][:3])}")
                    extra_str = f" ({', '.join(extras)})" if extras else ""
                    print(f"  [{col:2d}] {header}{extra_str}")

        # Print columns with variance
        mixed_cols = [(col, info) for col, info in dag.items()
                      if info["variance_pct"] > 0 or info["unique_formulas"] > 1]
        if mixed_cols:
            print("\n" + "=" * 80)
            print("COLUMNS WITH FORMULA VARIANCE")
            print("=" * 80)
            for col, info in sorted(mixed_cols):
                print(f"\n[Col {col}] {info['header']}")
                print(f"  Variance: {info['variance_pct']}% ({info['unique_formulas']} unique formulas)")
                print(f"  Base ({info['formula_rows'] - (info['formula_rows'] * info['variance_pct'] // 100)} rows): {info['base_formula'][:80] if info['base_formula'] else 'N/A'}...")

    # Print dependency tree
    print("\n" + "=" * 80)
    print("DEPENDENCY TREE")
    print("=" * 80)

    if args.target:
        target_col = args.target
    else:
        # Find the column with highest level
        level_cache = {}
        max_level = -1
        target_col = target_columns[0] if target_columns else 1
        for col in dag:
            level = get_level(col, dag, level_cache)
            if level > max_level:
                max_level = level
                target_col = col

    print(f"\nStarting from column {target_col} ({col_to_header.get(target_col, 'Unknown')}):\n")
    print_tree(target_col, dag, col_to_header)

    if not args.quiet:
        # Group by dependency level
        print("\n" + "=" * 80)
        print("CALCULATION LEVELS")
        print("=" * 80)

        levels = defaultdict(list)
        level_cache = {}
        for col in dag:
            level = get_level(col, dag, level_cache)
            levels[level].append(col)

        for level in sorted(levels.keys()):
            cols = levels[level]
            print(f"\nLevel {level}:")
            for col in sorted(cols):
                info = dag[col]
                header = info["header"]
                cls = info["classification"]
                external = info.get("external_deps", [])
                ext_str = f" → {', '.join(external)}" if external else ""
                print(f"  [{col:2d}] {header} ({cls}){ext_str}")


if __name__ == "__main__":
    main()
