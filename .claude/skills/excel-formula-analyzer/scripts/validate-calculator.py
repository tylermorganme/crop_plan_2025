#!/usr/bin/env python3
"""
Template script for validating a code implementation against Excel calculations.

This script compares calculated values from your code implementation against
the actual values from an Excel spreadsheet to ensure accuracy.

Usage:
    python validate-calculator.py <data.json> <workbook.xlsx> <sheet_name>

The data.json file should contain your calculated values.
The workbook should contain the "expected" values from Excel.

Customize the compare_values() function for your specific use case.
"""

import json
import sys
from datetime import datetime, timedelta

import openpyxl


def excel_to_date(excel_num):
    """Convert Excel date number to Python date."""
    if excel_num is None or excel_num < 1:
        return None
    # Excel's epoch is 1899-12-30 (accounting for the 1900 leap year bug)
    return datetime(1899, 12, 30) + timedelta(days=excel_num)


def format_date(d):
    """Format a date for display."""
    if d is None:
        return "null"
    if isinstance(d, str):
        return d[:10]  # Just the date part
    if isinstance(d, datetime):
        return d.strftime("%Y-%m-%d")
    return str(d)


def days_diff(d1, d2):
    """Calculate the difference in days between two dates."""
    if d1 is None or d2 is None:
        return None

    if isinstance(d1, str):
        d1 = datetime.fromisoformat(d1[:10])
    if isinstance(d2, str):
        d2 = datetime.fromisoformat(d2[:10])

    return (d2 - d1).days


def load_excel_values(workbook_path, sheet_name, header_row=5, data_start_row=6):
    """Load values from Excel for comparison."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]

    # Build header map
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[val] = col

    # Load data rows
    data = []
    for row in range(data_start_row, ws.max_row + 1):
        # Check if row has data (using first column as indicator)
        if ws.cell(row=row, column=1).value is None:
            continue

        row_data = {}
        for header, col in headers.items():
            value = ws.cell(row=row, column=col).value
            row_data[header] = value

        data.append(row_data)

    return data


def compare_values(calculated, expected, field_name):
    """
    Compare a calculated value against the expected Excel value.

    Returns: (match, diff, message)
        match: bool - True if values match
        diff: number or None - difference if applicable
        message: str - description of the comparison
    """
    if calculated is None and expected is None:
        return True, None, "Both null"

    if calculated is None:
        return False, None, f"Calculated is null, expected {expected}"

    if expected is None:
        return False, None, f"Expected is null, calculated {calculated}"

    # For dates, compare the values
    if "date" in field_name.lower():
        calc_str = format_date(calculated)
        exp_str = format_date(expected)

        if calc_str == exp_str:
            return True, 0, "Exact match"

        diff = days_diff(expected, calculated)
        if diff is not None and abs(diff) <= 1:
            return True, diff, f"Close match (+/- 1 day)"

        return False, diff, f"Mismatch: calc={calc_str}, exp={exp_str}"

    # For numbers, check within tolerance
    if isinstance(calculated, (int, float)) and isinstance(expected, (int, float)):
        if abs(calculated - expected) < 0.001:
            return True, 0, "Exact match"
        return False, calculated - expected, f"Mismatch: calc={calculated}, exp={expected}"

    # String comparison
    if str(calculated) == str(expected):
        return True, None, "Exact match"

    return False, None, f"Mismatch: calc={calculated}, exp={expected}"


def main():
    if len(sys.argv) < 4:
        print("Usage: python validate-calculator.py <data.json> <workbook.xlsx> <sheet_name>")
        print("\nExample:")
        print('  python validate-calculator.py calculated.json "Crop Plan.xlsx" "Bed Plan"')
        sys.exit(1)

    json_path = sys.argv[1]
    workbook_path = sys.argv[2]
    sheet_name = sys.argv[3]

    # Load calculated data
    print(f"Loading calculated data from: {json_path}")
    with open(json_path, "r") as f:
        calculated_data = json.load(f)

    # Load Excel data
    print(f"Loading Excel data from: {workbook_path} / {sheet_name}")
    excel_data = load_excel_values(workbook_path, sheet_name)

    print(f"\nComparing {len(calculated_data)} calculated records...")
    print("=" * 80)

    # Build lookup from Excel data
    # Customize this key based on your data's unique identifier
    excel_lookup = {row.get("Identifier"): row for row in excel_data}

    # Track results
    exact_matches = 0
    close_matches = 0
    mismatches = []
    skipped = 0

    # Fields to compare - customize for your use case
    fields_to_compare = [
        "End of Harvest",
        "Expected End of Harvest",
        "Beginning of Harvest",
        "Start Date",
    ]

    for calc_record in calculated_data:
        identifier = calc_record.get("identifier")
        if not identifier:
            skipped += 1
            continue

        excel_record = excel_lookup.get(identifier)
        if not excel_record:
            skipped += 1
            continue

        record_matches = True
        record_diffs = []

        for field in fields_to_compare:
            calc_value = calc_record.get(field.lower().replace(" ", "_"))
            exp_value = excel_record.get(field)

            match, diff, message = compare_values(calc_value, exp_value, field)

            if not match:
                record_matches = False
                record_diffs.append(f"{field}: {message}")
            elif diff is not None and diff != 0:
                record_diffs.append(f"{field}: {message}")

        if record_matches:
            if any("Close" in d for d in record_diffs):
                close_matches += 1
            else:
                exact_matches += 1
        else:
            mismatches.append({
                "identifier": identifier,
                "differences": record_diffs
            })

    # Print summary
    total_compared = exact_matches + close_matches + len(mismatches)

    print("\nSUMMARY")
    print("-" * 50)
    print(f"Total records: {len(calculated_data)}")
    print(f"Skipped (no match in Excel): {skipped}")
    print(f"Compared: {total_compared}")
    print()
    print(f"  Exact matches: {exact_matches}")
    print(f"  Close matches: {close_matches}")
    print(f"  Mismatches: {len(mismatches)}")

    if mismatches:
        print("\nMISMATCHES (first 20)")
        print("-" * 80)
        for m in mismatches[:20]:
            print(f"\n{m['identifier']}:")
            for diff in m["differences"]:
                print(f"  - {diff}")

    # Return exit code
    if len(mismatches) == 0:
        print("\n[OK] All values match!")
        sys.exit(0)
    else:
        print(f"\n[WARN] {len(mismatches)} mismatches found")
        sys.exit(1)


if __name__ == "__main__":
    main()
