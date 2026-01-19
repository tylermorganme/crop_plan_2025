#!/usr/bin/env python3
"""
Extract Products table from Excel workbook.

Usage:
    python extract-products.py
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


def main():
    # Find the workbook
    project_root = Path(__file__).parent.parent.parent
    workbook_path = project_root / "Crop Plan 2025 V20.xlsm"

    if not workbook_path.exists():
        print(f"Error: Workbook not found at {workbook_path}")
        sys.exit(1)

    print(f"Loading workbook: {workbook_path}")
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)

    # List all sheets
    print(f"\nSheets in workbook ({len(wb.sheetnames)}):")
    for i, name in enumerate(wb.sheetnames):
        print(f"  {i+1}. {name}")

    # Look for Products sheet
    products_sheet = None
    for name in wb.sheetnames:
        if 'product' in name.lower():
            products_sheet = name
            break

    if not products_sheet:
        print("\nNo 'Products' sheet found. Looking for tables...")
        # Check each sheet for tables
        wb_full = openpyxl.load_workbook(workbook_path, data_only=True)
        for sheet_name in wb_full.sheetnames:
            ws = wb_full[sheet_name]
            if hasattr(ws, 'tables') and ws.tables:
                print(f"  {sheet_name}: {list(ws.tables.keys())}")
        wb_full.close()
        return

    print(f"\nExtracting from sheet: {products_sheet}")
    ws = wb[products_sheet]

    # Get all rows first to find where actual data starts
    all_rows = list(ws.iter_rows(values_only=True))

    # Find the header row (look for "ID" or "Crop" in first few columns)
    header_row_idx = None
    for i, row in enumerate(all_rows[:5]):  # Check first 5 rows
        if row and any(cell in ['ID', 'Crop', 'Product'] for cell in row[:5] if cell):
            header_row_idx = i
            break

    if header_row_idx is None:
        print("Could not find header row")
        return

    print(f"Header row found at index {header_row_idx}")

    headers = []
    data = []

    # Get headers
    header_row = all_rows[header_row_idx]
    headers = [str(cell).strip() if cell else f"col_{i}" for i, cell in enumerate(header_row)]

    # Get data rows
    for row in all_rows[header_row_idx + 1:]:
        # Skip empty rows
        if not any(row):
            continue

        # Skip rows where first cell is empty or looks like a formula result
        first_cell = row[0]
        if first_cell is None:
            continue

        record = {}
        for i, value in enumerate(row):
            if i < len(headers):
                record[headers[i]] = value
        data.append(record)

    print(f"Found {len(data)} products")
    print(f"Headers: {headers}")

    # Save to JSON
    output_path = project_root / "crop-api" / "src" / "data" / "products-template.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump({
            "extractedAt": str(Path(workbook_path).stat().st_mtime),
            "sourceSheet": products_sheet,
            "headers": headers,
            "products": data
        }, f, indent=2, default=str)

    print(f"\nSaved to: {output_path}")

    # Print sample
    if data:
        print("\nSample product:")
        print(json.dumps(data[0], indent=2, default=str))

    wb.close()


if __name__ == "__main__":
    main()
