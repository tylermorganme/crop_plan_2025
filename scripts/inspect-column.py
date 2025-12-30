#!/usr/bin/env python3
"""
Inspect a column from the Crop Chart or Bed Plan.

Usage:
  python scripts/inspect-column.py "STH"
  python scripts/inspect-column.py "DTM" --sheet "Bed Plan"
  python scripts/inspect-column.py --list              # List all columns
  python scripts/inspect-column.py --list --sheet "Bed Plan"
"""

import openpyxl
import argparse
import sys

def get_column_info(ws, header_row=2):
    """Get all column headers and their positions."""
    columns = {}
    for col in range(1, 200):
        val = ws.cell(row=header_row, column=col).value
        if val:
            columns[val] = col
        # Also check row above for multi-row headers
        val_above = ws.cell(row=header_row-1, column=col).value
        if val_above and val_above not in columns:
            columns[val_above] = col
    return columns

def inspect_column(ws, col_num, header, num_rows=20, header_row=2):
    """Inspect a column's formulas and values."""
    print(f"\n{'='*70}")
    print(f"Column {col_num}: {header}")
    print('='*70)

    # Check first data row for formula
    first_data_row = header_row + 1
    cell = ws.cell(row=first_data_row, column=col_num)
    formula = cell.value

    if formula and str(formula).startswith('='):
        print(f"\nFORMULA (row {first_data_row}):")
        print(f"  {formula}")

        # Check if formula varies across rows
        formulas_vary = False
        for row in range(first_data_row + 1, min(first_data_row + 5, ws.max_row)):
            other_formula = ws.cell(row=row, column=col_num).value
            if other_formula != formula:
                formulas_vary = True
                break

        if formulas_vary:
            print("\n  [!] Formula varies across rows")
        else:
            print("\n  [OK] Formula is consistent across rows")
    else:
        print(f"\nNO FORMULA - appears to be static input")
        print(f"  First value: {formula}")

    # Show sample data
    print(f"\nSAMPLE DATA (rows {first_data_row}-{first_data_row + num_rows - 1}):")
    print("-"*70)

    for row in range(first_data_row, min(first_data_row + num_rows, ws.max_row)):
        val = ws.cell(row=row, column=col_num).value
        val_str = str(val)[:60] if val else "(empty)"
        print(f"  Row {row:3d}: {val_str}")

def list_columns(ws, header_row=2):
    """List all columns with their types."""
    columns = get_column_info(ws, header_row)

    print(f"\n{'='*70}")
    print(f"ALL COLUMNS (header row {header_row})")
    print('='*70)

    # Sort by column number
    sorted_cols = sorted(columns.items(), key=lambda x: x[1])

    for header, col_num in sorted_cols:
        cell = ws.cell(row=header_row + 1, column=col_num)
        val = cell.value

        if val and str(val).startswith('='):
            col_type = "FORMULA"
            # Check for array formula marker
            if 'openpyxl.worksheet.formula' in str(type(val)):
                col_type = "ARRAY_FORMULA"
        elif val is None or val == "":
            col_type = "EMPTY"
        else:
            col_type = "STATIC"

        print(f"  {col_num:3d} | {header[:35]:35s} | {col_type}")

def main():
    parser = argparse.ArgumentParser(description='Inspect Excel columns')
    parser.add_argument('column', nargs='?', help='Column header to inspect')
    parser.add_argument('--sheet', default='Crop Chart', help='Sheet name')
    parser.add_argument('--list', action='store_true', help='List all columns')
    parser.add_argument('--rows', type=int, default=15, help='Number of sample rows')
    parser.add_argument('--workbook', default='Crop Plan 2025 V20.xlsm', help='Workbook path')

    args = parser.parse_args()

    print(f"Loading {args.workbook}...")
    wb = openpyxl.load_workbook(args.workbook, data_only=False)

    if args.sheet not in wb.sheetnames:
        print(f"Error: Sheet '{args.sheet}' not found")
        print(f"Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[args.sheet]

    # Determine header row (Crop Chart uses row 2, Bed Plan uses row 5)
    header_row = 5 if args.sheet == 'Bed Plan' else 2

    if args.list:
        list_columns(ws, header_row)
        return

    if not args.column:
        print("Error: Please specify a column name or use --list")
        sys.exit(1)

    columns = get_column_info(ws, header_row)

    # Find the column (case-insensitive partial match)
    matches = [(h, c) for h, c in columns.items()
               if args.column.lower() in h.lower()]

    if not matches:
        print(f"Error: Column '{args.column}' not found")
        print(f"Available columns: {list(columns.keys())[:20]}...")
        sys.exit(1)

    if len(matches) > 1:
        print(f"Multiple matches for '{args.column}':")
        for h, c in matches:
            print(f"  {c}: {h}")
        print("\nInspecting first match...")

    header, col_num = matches[0]
    inspect_column(ws, col_num, header, args.rows, header_row)

if __name__ == '__main__':
    main()
