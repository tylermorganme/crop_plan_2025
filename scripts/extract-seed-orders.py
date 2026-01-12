#!/usr/bin/env python3
"""
Extract seed order data from Excel Seed List sheet.

This script:
1. Reads seed order data from the Excel Seed List sheet
2. Matches varieties by crop|name|supplier key
3. Outputs seed-orders.json for import into the app
"""

import openpyxl
import json
from datetime import datetime, timezone

# Unit mapping from Excel to our ProductUnit type
UNIT_MAP = {
    'g': 'g',
    'ozm': 'oz',
    'oz': 'oz',
    'lbm': 'lb',
    'lb': 'lb',
    'ct': 'ct',
}


def extract_seed_orders(excel_path: str) -> list:
    """Extract seed order data from Excel Seed List sheet."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Seed List']

    orders = []

    for row in range(6, 265):  # Seed List data starts at row 6
        crop = ws.cell(row=row, column=11).value  # Crop
        company = ws.cell(row=row, column=12).value  # Company
        variety = ws.cell(row=row, column=13).value  # Variety

        if not crop or not variety or not company:
            continue

        order_flag = ws.cell(row=row, column=14).value  # Order
        already_have = ws.cell(row=row, column=24).value  # Already Have

        # Skip if not ordering and don't already have
        if not order_flag and not already_have:
            continue

        prod_weight = ws.cell(row=row, column=20).value  # Product Weight
        weight_unit = ws.cell(row=row, column=21).value  # Weight Unit
        prod_cost = ws.cell(row=row, column=22).value  # Product Cost
        prod_qty = ws.cell(row=row, column=23).value  # Product Quantity
        prod_link = ws.cell(row=row, column=18).value  # Link
        notes = ws.cell(row=row, column=31).value  # Notes

        # Create variety key (matches our ID format)
        variety_key = f'{crop.lower().strip()}|{variety.lower().strip()}|{company.lower().strip()}'
        variety_id = f'V_{variety_key}'

        # Map unit
        product_unit = UNIT_MAP.get(weight_unit) if weight_unit else None

        order = {
            'varietyId': variety_id,
            '_crop': crop,
            '_variety': variety,
            '_company': company,
        }

        if prod_weight is not None and prod_weight != '':
            order['productWeight'] = float(prod_weight) if prod_weight else None
        if product_unit:
            order['productUnit'] = product_unit
        if prod_cost is not None and prod_cost != '':
            order['productCost'] = float(prod_cost) if prod_cost else None
        if prod_qty is not None and prod_qty != '':
            order['quantity'] = int(prod_qty) if prod_qty else 0
        else:
            order['quantity'] = 0
        if already_have:
            order['alreadyHave'] = True
        if prod_link and str(prod_link).lower() != 'link' and str(prod_link).startswith('http'):
            order['productLink'] = str(prod_link)
        if notes:
            order['notes'] = str(notes)

        orders.append(order)

    return orders


def main():
    excel_path = 'Crop Plan 2025 V20.xlsm'
    output_path = 'src/data/seed-orders.json'

    print(f'Extracting seed orders from {excel_path}...')
    orders = extract_seed_orders(excel_path)
    print(f'Found {len(orders)} seed orders')

    # Count stats
    with_product = sum(1 for o in orders if 'productWeight' in o)
    already_have = sum(1 for o in orders if o.get('alreadyHave'))

    print(f'  With product info: {with_product}')
    print(f'  Already have: {already_have}')

    # Output
    output_data = {
        '_generated': datetime.now(timezone.utc).isoformat(),
        '_source': excel_path,
        'seedOrders': orders,
    }

    with open(output_path, 'w') as f:
        json.dump(output_data, f, indent=2)

    print(f'\nWrote {output_path}')

    # Show samples
    print('\nSample orders:')
    for order in orders[:5]:
        print(f"  {order['_crop']} - {order['_variety']} ({order['_company']})")
        if 'productWeight' in order:
            print(f"    Product: {order.get('productWeight')} {order.get('productUnit')} @ ${order.get('productCost')} x {order.get('quantity')}")
        if order.get('alreadyHave'):
            print(f"    Already have: True")


if __name__ == '__main__':
    main()
