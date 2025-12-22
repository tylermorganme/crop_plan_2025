"""
Parse structured table references in Bed Plan formulas and build a dependency DAG.
Structured refs like BedPlan[[#This Row],[Column Name]] resolve to same-row column lookups.
"""

import openpyxl
import re
from collections import defaultdict

# Load workbook WITH formulas (not data_only)
print("Loading workbook...")
wb = openpyxl.load_workbook("Crop Plan 2025 V20.xlsm", data_only=False)
ws = wb["Bed Plan"]

# Build header -> column mapping
headers = {}
col_to_header = {}
for col in range(1, 160):
    val = ws.cell(row=5, column=col).value
    if val:
        headers[val] = col
        col_to_header[col] = val

print(f"Found {len(headers)} headers")

def parse_structured_refs(formula):
    """
    Extract column references from structured table references.
    Pattern: BedPlan[[#This Row],[Column Name]] -> Column Name
    Also handles: BedPlan[Column Name] for full column refs
    """
    if not formula or not str(formula).startswith("="):
        return []

    refs = []

    # Pattern for same-row references: BedPlan[[#This Row],[Column Name]]
    same_row = re.findall(r"BedPlan\[\[#This Row\],\[([^\]]+)\]\]", str(formula))
    refs.extend(same_row)

    # Pattern for XLOOKUP referencing other rows: BedPlan[Column Name]
    # These reference entire columns, not specific cells
    col_refs = re.findall(r"BedPlan\[([^\[\]]+)\]", str(formula))
    refs.extend(col_refs)

    return list(set(refs))

def get_formula_dependencies(col_num, row=6):
    """Get the column dependencies for a formula in a given column."""
    cell = ws.cell(row=row, column=col_num)
    formula = cell.value

    if not formula or not str(formula).startswith("="):
        return None, []

    col_refs = parse_structured_refs(str(formula))

    # Convert column names to column numbers
    deps = []
    for ref in col_refs:
        if ref in headers:
            deps.append(headers[ref])

    return str(formula), deps

def build_dag():
    """Build the complete dependency DAG for timing columns."""
    dag = {}

    # Key timing columns
    timing_cols = [16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 35, 36]
    # Config columns that might be referenced
    config_cols = [57, 58, 59, 60, 62, 71, 72]

    all_cols = timing_cols + config_cols

    for col in all_cols:
        header = col_to_header.get(col, f"Col {col}")
        formula, deps = get_formula_dependencies(col)

        dag[col] = {
            "header": header,
            "formula": formula,
            "depends_on": deps
        }

    return dag

# Build the DAG
print("\nBuilding dependency DAG...")
dag = build_dag()

# Print the DAG
print("\n" + "="*80)
print("FORMULA DEPENDENCY DAG - Bed Plan Sheet (Row 6)")
print("="*80)

def get_dep_chain(col, dag, visited=None):
    """Recursively get all dependencies"""
    if visited is None:
        visited = set()
    if col in visited:
        return []
    visited.add(col)

    direct_deps = dag.get(col, {}).get("depends_on", [])
    all_deps = list(direct_deps)

    for dep in direct_deps:
        all_deps.extend(get_dep_chain(dep, dag, visited))

    return all_deps

# Print each node
for col in sorted(dag.keys()):
    info = dag[col]
    header = info["header"]
    formula = info["formula"]
    deps = info["depends_on"]

    print(f"\n[Col {col:2d}] {header}")

    if formula:
        # Truncate long formulas
        if len(formula) > 100:
            formula = formula[:100] + "..."
        print(f"  Formula: {formula}")

        if deps:
            dep_names = [col_to_header.get(d, f"Col {d}") for d in deps]
            print(f"  Depends on: {dep_names}")
        else:
            print(f"  Depends on: (no table references - may have cell refs or constants)")
    else:
        print(f"  Type: INPUT (no formula)")

# Now trace from End of Harvest backwards
print("\n\n" + "="*80)
print("DEPENDENCY CHAIN: End of Harvest -> Start Date")
print("="*80)

def print_tree(col, dag, indent=0, visited=None):
    """Print dependency tree"""
    if visited is None:
        visited = set()
    if col in visited:
        return
    visited.add(col)

    info = dag.get(col, {})
    header = info.get("header", f"Col {col}")
    formula = info.get("formula")
    deps = info.get("depends_on", [])

    prefix = "  " * indent + ("|- " if indent > 0 else "")

    if formula:
        formula_short = formula[:70] + "..." if len(str(formula)) > 70 else formula
        print(f"{prefix}[{col}] {header}")
        print(f"{prefix}     = {formula_short}")
    else:
        print(f"{prefix}[{col}] {header} (INPUT)")

    for dep in deps:
        print_tree(dep, dag, indent + 1, visited)

# Start from End of Harvest (col 36)
print("\nStarting from End of Harvest (col 36):\n")
print_tree(36, dag)

# Also show the flow in reverse (inputs to outputs)
print("\n\n" + "="*80)
print("CALCULATION FLOW (Forward)")
print("="*80)

# Group by dependency level
levels = defaultdict(list)

def get_level(col, dag, cache=None):
    if cache is None:
        cache = {}
    if col in cache:
        return cache[col]

    deps = dag.get(col, {}).get("depends_on", [])
    if not deps:
        cache[col] = 0
        return 0

    max_dep_level = max(get_level(d, dag, cache) for d in deps)
    cache[col] = max_dep_level + 1
    return cache[col]

level_cache = {}
for col in dag:
    level = get_level(col, dag, level_cache)
    levels[level].append(col)

for level in sorted(levels.keys()):
    cols = levels[level]
    print(f"\nLevel {level}:")
    for col in cols:
        header = dag[col]["header"]
        formula = dag[col].get("formula")
        if formula:
            print(f"  [{col:2d}] {header} (calculated)")
        else:
            print(f"  [{col:2d}] {header} (INPUT)")

# Summary
print("\n\n" + "="*80)
print("SUMMARY: Key Calculation Path")
print("="*80)

summary = """
INPUTS -> End of Harvest:

Level 0 (INPUTS - User enters these):
  - [20] Fixed Field Start Date
  - [21] Follows Crop
  - [22] Follow Offset
  - [18] Actual Greenhouse Date
  - [24] Actual TP or DS Date
  - [29] Actual Beginning of Harvest
  - [32] Actual End of Harvest
  - [33] Additional Days of Harvest
  - [35] Failed
  - [57] DTM
  - [59] Harvest Window
  - [62] Days in Cells

Level 1 (First calculations):
  - [17] Planned Greenhouse Start Date = f(Fixed Field Start Date, Days in Cells, Follows Crop...)
  - [23] Planned TP or DS Date = f(Actual Greenhouse Date, Days in Cells, Follows Crop...)

Level 2:
  - [19] Greenhouse Start Date = COALESCE(Actual, Planned)
  - [26] TP or DS Date = COALESCE(Actual, Planned)
  - [16] Start Date = Earlier of GH Start or Field Date

Level 3:
  - [28] Expected Beginning of Harvest = GH Start + DTM or TP/DS + DTM

Level 4:
  - [30] Beginning of Harvest = COALESCE(Actual, Expected)

Level 5:
  - [31] Expected End of Harvest = Expected Beginning + Harvest Window + Additional Days

Level 6 (FINAL OUTPUT):
  - [36] End of Harvest = COALESCE(Actual End, Expected End)
"""
print(summary)
