#!/usr/bin/env python3
"""
Extract variety data from the Varieties sheet in the Excel workbook.
Creates a JSON file for the variety import pipeline.

Usage:
    python scripts/extract-varieties.py
"""

import openpyxl
import json
from datetime import datetime, date
from pathlib import Path

# Custom JSON encoder for datetime objects
class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)

def fmt_value(val):
    """Format a cell value for JSON output"""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip() if val.strip() else None
    return val

def to_bool(val):
    """Convert Excel boolean/string to Python bool"""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.lower() in ('true', 'yes', '1', 'x')
    return bool(val)

# Load the workbook
print("Loading workbook...")
wb = openpyxl.load_workbook("Crop Plan 2025 V20.xlsm", data_only=True)

# Check available sheets
print(f"Available sheets: {wb.sheetnames}")

# Read the Varieties sheet
if "Varieties" not in wb.sheetnames:
    print("ERROR: 'Varieties' sheet not found!")
    exit(1)

ws = wb["Varieties"]

# Get headers from row 2 (row 1 is empty in this sheet)
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=2, column=col).value
    headers.append(header)

print(f"Found {len(headers)} columns: {headers}")

# Expected column mapping (adjust based on actual sheet structure)
# Common columns: Id, Crop, Sub Category, Variety, DTM, Company, Organic, Pelleted, etc.

# Extract variety data starting from row 3
varieties = []
for row in range(3, ws.max_row + 1):
    row_data = {}
    for col, header in enumerate(headers, start=1):
        if header:
            value = ws.cell(row=row, column=col).value
            row_data[header] = fmt_value(value)

    # Skip empty rows
    if not row_data.get("Crop") and not row_data.get("Variety"):
        continue

    # Transform to our variety format
    variety = {
        "crop": row_data.get("Crop") or "",
        "name": row_data.get("Variety") or row_data.get("Name") or "",
        "supplier": row_data.get("Company") or row_data.get("Supplier") or "",
        "organic": to_bool(row_data.get("Organic")),
        "pelleted": to_bool(row_data.get("Pelleted")),
        "pelletedApproved": to_bool(row_data.get("Approved Pellet") or row_data.get("Pelleted Approved")),
        "dtm": row_data.get("DTM"),
        "subCategory": row_data.get("Sub Category"),
        "website": row_data.get("Website") or row_data.get("URL"),
        "alreadyOwn": to_bool(row_data.get("Already Own") or row_data.get("Have")),
        # Keep original ID for reference
        "excelId": row_data.get("Id"),
    }

    # Filter out None values
    variety = {k: v for k, v in variety.items() if v is not None and v != ""}

    if variety.get("crop") or variety.get("name"):
        varieties.append(variety)

print(f"Extracted {len(varieties)} varieties")

# Ensure output directory exists
output_dir = Path("tmp")
output_dir.mkdir(exist_ok=True)

# Save to JSON
output_path = output_dir / "varieties_from_excel.json"
with open(output_path, "w") as f:
    json.dump({"varieties": varieties}, f, indent=2, cls=DateTimeEncoder)

print(f"Saved to {output_path}")

# Show sample and stats
print("\n--- Statistics ---")
crops = set(v.get("crop", "") for v in varieties if v.get("crop"))
suppliers = set(v.get("supplier", "") for v in varieties if v.get("supplier"))
organic_count = sum(1 for v in varieties if v.get("organic"))

print(f"Unique crops: {len(crops)}")
print(f"Unique suppliers: {len(suppliers)}")
print(f"Organic varieties: {organic_count}")

print("\nSample variety:")
if varieties:
    print(json.dumps(varieties[0], indent=2))
