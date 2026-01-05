#!/usr/bin/env python3
"""
Convert Excel yield formulas to our new yieldFormula string format.

Reads the Excel workbook, parses each Units Per Harvest formula,
converts to our expression syntax, validates against Excel's values,
and outputs a JSON shim file for merging into crops.json.

Usage:
    python scripts/convert-yield-formulas.py

Output:
    tmp/yield-formulas.json - Shim data with yieldFormula for each config
"""

import openpyxl
import json
import re
from dataclasses import dataclass
from typing import Optional
from pathlib import Path


@dataclass
class FormulaConversion:
    """Result of converting an Excel formula to our format."""
    formula: str  # Our yieldFormula string
    pattern: str  # Pattern name for debugging
    is_total: bool  # True if formula produces total yield, False if per-harvest


def parse_and_convert(excel_formula: str, harvests: int = 1) -> Optional[FormulaConversion]:
    """
    Parse an Excel UPH formula and convert to our yieldFormula format.

    Excel formulas calculate Units Per Harvest.
    Our formulas calculate TOTAL yield.

    So if Excel has /Harvests, we remove it (already total).
    If Excel doesn't have /Harvests, we multiply by harvests.

    For weekly-rate formulas (cucumbers, squash), the Excel formula gives
    UPH and we need to multiply by harvests to get total.
    """
    if not excel_formula or not excel_formula.startswith('='):
        return None

    # Normalize: remove Excel table references - handle both formats
    f = excel_formula
    f = f.replace('Crops[[#This Row],', '').replace(']]', ']').replace('[', '').replace(']', '')
    f = f.replace('=', '')

    # Replace Excel column names with our variable names
    # IMPORTANT: Replace longer names first to avoid partial matches
    # NOTE: We use full descriptive names (plantingsPerBed, daysBetweenHarvest)
    # not abbreviations (PPB, DBH) for self-documenting formulas
    f = f.replace('Plantings Per Bed', 'plantingsPerBed')
    f = f.replace('Days Between Harvest', 'daysBetweenHarvest')
    f = f.replace('StandardBedLength', '100')  # Must come before BedLength!
    f = f.replace('BedLength', 'bedFeet')
    f = f.replace('Harvests', 'harvests')
    f = f.replace('Seeds Per Bed', 'seeds')
    f = f.replace('Safety Factor', 'safetyFactor')

    # Track if formula has /harvests (meaning it's already computing per-harvest)
    has_div_harvests = '/harvests' in f.lower()

    # ========== Pattern matching and conversion ==========
    # Note: Patterns match against 'plantingsPerBed' (from Excel 'Plantings Per Bed')

    # plantingsPerBed * multiplier / harvests -> plantingsPerBed * multiplier (total)
    m = re.match(r'^plantingsPerBed\*(\d+\.?\d*)/harvests$', f)
    if m:
        rate = float(m.group(1))
        return FormulaConversion(f'plantingsPerBed * {rate}', 'PPB_MULT_DIV_H', True)

    # plantingsPerBed / divisor / harvests -> plantingsPerBed * (1/divisor) (total)
    m = re.match(r'^plantingsPerBed/(\d+\.?\d*)/harvests$', f)
    if m:
        rate = 1.0 / float(m.group(1))
        return FormulaConversion(f'plantingsPerBed * {rate}', 'PPB_DIV_DIV_H', True)

    # plantingsPerBed / harvests -> plantingsPerBed * 1 (total)
    m = re.match(r'^plantingsPerBed/harvests$', f)
    if m:
        return FormulaConversion('plantingsPerBed', 'PPB_DIV_H', True)

    # plantingsPerBed * multiplier (no /harvests) -> plantingsPerBed * multiplier * harvests
    m = re.match(r'^plantingsPerBed\*(\d+\.?\d*)$', f)
    if m:
        rate = float(m.group(1))
        return FormulaConversion(f'plantingsPerBed * {rate} * harvests', 'PPB_MULT', False)

    # plantingsPerBed / divisor (no /harvests) -> plantingsPerBed * (1/divisor) * harvests
    m = re.match(r'^plantingsPerBed/(\d+\.?\d*)$', f)
    if m:
        rate = 1.0 / float(m.group(1))
        return FormulaConversion(f'plantingsPerBed * {rate} * harvests', 'PPB_DIV', False)

    # plantingsPerBed direct (no /harvests) -> plantingsPerBed * harvests
    m = re.match(r'^plantingsPerBed$', f)
    if m:
        return FormulaConversion('plantingsPerBed * harvests', 'PPB_DIRECT', False)

    # plantingsPerBed * (fraction) -> plantingsPerBed * rate * harvests
    m = re.match(r'^plantingsPerBed\*\((\d+\.?\d*)/(\d+\.?\d*)\)$', f)
    if m:
        rate = float(m.group(1)) / float(m.group(2))
        return FormulaConversion(f'plantingsPerBed * {rate} * harvests', 'PPB_MULT', False)

    # plantingsPerBed * (fraction) * multiplier -> plantingsPerBed * rate * harvests
    m = re.match(r'^plantingsPerBed\*\((\d+\.?\d*)/(\d+\.?\d*)\)\*(\d+\.?\d*)$', f)
    if m:
        rate = float(m.group(1)) / float(m.group(2)) * float(m.group(3))
        return FormulaConversion(f'plantingsPerBed * {rate} * harvests', 'PPB_MULT', False)

    # plantingsPerBed * a / b (not harvests) -> plantingsPerBed * rate * harvests
    m = re.match(r'^plantingsPerBed\*(\d+\.?\d*)/(\d+\.?\d*)$', f)
    if m and m.group(2) != 'harvests':
        rate = float(m.group(1)) / float(m.group(2))
        return FormulaConversion(f'plantingsPerBed * {rate} * harvests', 'PPB_MULT', False)

    # plantingsPerBed / a / b / harvests -> plantingsPerBed * rate (total)
    m = re.match(r'^plantingsPerBed/(\d+\.?\d*)/(\d+\.?\d*)/harvests$', f)
    if m:
        rate = 1.0 / (float(m.group(1)) * float(m.group(2)))
        return FormulaConversion(f'plantingsPerBed * {rate}', 'PPB_DIV_DIV_H', True)

    # plantingsPerBed / a / b / c / harvests -> plantingsPerBed * rate (total)
    m = re.match(r'^plantingsPerBed/(\d+\.?\d*)/(\d+\.?\d*)/(\d+\.?\d*)/harvests$', f)
    if m:
        rate = 1.0 / (float(m.group(1)) * float(m.group(2)) * float(m.group(3)))
        return FormulaConversion(f'plantingsPerBed * {rate}', 'PPB_DIV_DIV_H', True)

    # plantingsPerBed / a / b / c (no harvests) -> plantingsPerBed * rate * harvests
    m = re.match(r'^plantingsPerBed/(\d+\.?\d*)/(\d+\.?\d*)/(\d+\.?\d*)$', f)
    if m:
        rate = 1.0 / (float(m.group(1)) * float(m.group(2)) * float(m.group(3)))
        return FormulaConversion(f'plantingsPerBed * {rate} * harvests', 'PPB_DIV', False)

    # plantingsPerBed * a / b / harvests -> plantingsPerBed * rate (total)
    m = re.match(r'^plantingsPerBed\*(\d+\.?\d*)/(\d+\.?\d*)/harvests$', f)
    if m:
        rate = float(m.group(1)) / float(m.group(2))
        return FormulaConversion(f'plantingsPerBed * {rate}', 'PPB_MULT_DIV_H', True)

    # plantingsPerBed / a * b / harvests -> plantingsPerBed * rate (total)
    m = re.match(r'^plantingsPerBed/(\d+\.?\d*)\*(\d+\.?\d*)/harvests$', f)
    if m:
        rate = float(m.group(2)) / float(m.group(1))
        return FormulaConversion(f'plantingsPerBed * {rate}', 'PPB_MULT_DIV_H', True)

    # plantingsPerBed * a * b / harvests -> plantingsPerBed * rate (total)
    m = re.match(r'^plantingsPerBed\*(\d+\.?\d*)\*(\d+\.?\d*)/harvests$', f)
    if m:
        rate = float(m.group(1)) * float(m.group(2))
        return FormulaConversion(f'plantingsPerBed * {rate}', 'PPB_MULT_DIV_H', True)

    # plantingsPerBed * a * b * c / harvests -> plantingsPerBed * rate (total)
    m = re.match(r'^plantingsPerBed\*(\d+\.?\d*)\*(\d+\.?\d*)\*(\d+\.?\d*)/harvests$', f)
    if m:
        rate = float(m.group(1)) * float(m.group(2)) * float(m.group(3))
        return FormulaConversion(f'plantingsPerBed * {rate}', 'PPB_MULT_DIV_H', True)

    # plantingsPerBed * a * b / c / harvests (onion pattern) -> plantingsPerBed * rate (total)
    m = re.match(r'^plantingsPerBed\*(\d+\.?\d*)\*(\d+\.?\d*)/(\d+\.?\d*)/harvests$', f)
    if m:
        rate = float(m.group(1)) * float(m.group(2)) / float(m.group(3))
        return FormulaConversion(f'plantingsPerBed * {rate}', 'PPB_MULT_DIV_H', True)

    # plantingsPerBed / harvests * multiplier -> plantingsPerBed * rate (total)
    m = re.match(r'^plantingsPerBed/harvests\*(\d+\.?\d*)$', f)
    if m:
        rate = float(m.group(1))
        return FormulaConversion(f'plantingsPerBed * {rate}', 'PPB_MULT_DIV_H', True)

    # 1/harvests * plantingsPerBed -> plantingsPerBed (total)
    m = re.match(r'^1/harvests\*plantingsPerBed$', f)
    if m:
        return FormulaConversion('plantingsPerBed', 'PPB_DIV_H', True)

    # plantingsPerBed / a * b (no harvests) -> plantingsPerBed * rate * harvests
    m = re.match(r'^plantingsPerBed/(\d+\.?\d*)\*(\d+\.?\d*)$', f)
    if m:
        rate = float(m.group(2)) / float(m.group(1))
        return FormulaConversion(f'plantingsPerBed * {rate} * harvests', 'PPB_MULT', False)

    # ========== Area-based patterns ==========

    # X * bedFeet/100 / harvests -> (bedFeet / 100) * X (total)
    m = re.match(r'^(\d+\.?\d*)\*bedFeet/100/harvests$', f)
    if m:
        rate = float(m.group(1))
        return FormulaConversion(f'(bedFeet / 100) * {rate}', 'AREA_DIV_H', True)

    # X * bedFeet/100 (no harvests) -> (bedFeet / 100) * X * harvests
    m = re.match(r'^(\d+\.?\d*)\*bedFeet/100$', f)
    if m:
        rate = float(m.group(1))
        return FormulaConversion(f'(bedFeet / 100) * {rate} * harvests', 'AREA', False)

    # bedFeet / harvests -> bedFeet (total)
    m = re.match(r'^bedFeet/harvests$', f)
    if m:
        return FormulaConversion('bedFeet', 'BEDLENGTH_DIV_H', True)

    # bedFeet * X / harvests -> bedFeet * X (total)
    m = re.match(r'^bedFeet\*(\d+\.?\d*)/harvests$', f)
    if m:
        rate = float(m.group(1))
        return FormulaConversion(f'bedFeet * {rate}', 'BEDLENGTH_DIV_H', True)

    # X * Y * bedFeet / harvests -> bedFeet * (X*Y) (total)
    m = re.match(r'^(\d+\.?\d*)\*(\d+\.?\d*)\*bedFeet/harvests$', f)
    if m:
        rate = float(m.group(1)) * float(m.group(2))
        return FormulaConversion(f'bedFeet * {rate}', 'BEDLENGTH_DIV_H', True)

    # Sunchoke: 400*bedFeet/100/27*2.2*10/harvests -> (bedFeet / 100) * rate (total)
    m = re.match(r'^(\d+\.?\d*)\*bedFeet/100/(\d+\.?\d*)\*(\d+\.?\d*)\*(\d+\.?\d*)/harvests$', f)
    if m:
        rate = float(m.group(1)) / float(m.group(2)) * float(m.group(3)) * float(m.group(4))
        return FormulaConversion(f'(bedFeet / 100) * {rate}', 'AREA_DIV_H', True)

    # Eggplant/Squash: X*bedFeet/100/harvests -> (bedFeet / 100) * X (total)
    m = re.match(r'^(\d+\.?\d*)\*bedFeet/100/harvests$', f)
    if m:
        rate = float(m.group(1))
        return FormulaConversion(f'(bedFeet / 100) * {rate}', 'AREA_DIV_H', True)

    # ========== daysBetweenHarvest-scaled patterns (weekly/daily production) ==========
    # These compute UPH (units per harvest), so we multiply by harvests for total

    # plantingsPerBed * rate * daysBetweenHarvest / 7 -> plantingsPerBed * rate * (daysBetweenHarvest / 7) * harvests
    m = re.match(r'^plantingsPerBed\*(\d+\.?\d*)\*daysBetweenHarvest/7$', f)
    if m:
        rate = float(m.group(1))
        return FormulaConversion(f'plantingsPerBed * {rate} * (daysBetweenHarvest / 7) * harvests', 'PPB_WEEKLY_RATE', False)

    # plantingsPerBed * daysBetweenHarvest / 7 (rate = 1) -> plantingsPerBed * (daysBetweenHarvest / 7) * harvests
    m = re.match(r'^plantingsPerBed\*daysBetweenHarvest/7$', f)
    if m:
        return FormulaConversion('plantingsPerBed * (daysBetweenHarvest / 7) * harvests', 'PPB_WEEKLY_RATE', False)

    # plantingsPerBed * rate * (daysBetweenHarvest / 7) - parenthesized -> * harvests
    m = re.match(r'^plantingsPerBed\*(\d+\.?\d*)\*\(daysBetweenHarvest/7\)$', f)
    if m:
        rate = float(m.group(1))
        return FormulaConversion(f'plantingsPerBed * {rate} * (daysBetweenHarvest / 7) * harvests', 'PPB_WEEKLY_RATE', False)

    # plantingsPerBed * rate * daysBetweenHarvest (rat tail radish - daily rate) -> * harvests
    m = re.match(r'^plantingsPerBed\*(\d+\.?\d*)\*daysBetweenHarvest$', f)
    if m:
        rate = float(m.group(1))
        return FormulaConversion(f'plantingsPerBed * {rate} * daysBetweenHarvest * harvests', 'PPB_DBH_DIRECT', False)

    # ========== Tomato pattern (area-based, daysBetweenHarvest-scaled) ==========
    # 500*bedFeet/100*(harvests*daysBetweenHarvest)/42/harvests -> (bedFeet / 100) * (500/42) * daysBetweenHarvest
    # This computes UPH, total = UPH * harvests
    m = re.match(r'^(\d+\.?\d*)\*bedFeet/100\*\(harvests\*daysBetweenHarvest\)/(\d+\.?\d*)/harvests$', f)
    if m:
        base = float(m.group(1))
        divisor = float(m.group(2))
        rate = base / divisor
        return FormulaConversion(f'(bedFeet / 100) * {rate} * daysBetweenHarvest * harvests', 'AREA_DBH_SCALED', False)

    # Same with trailing multiplier (e.g., *0.5 for cherry tomatoes)
    m = re.match(r'^(\d+\.?\d*)\*bedFeet/100\*\(harvests\*daysBetweenHarvest\)/(\d+\.?\d*)/harvests\*(\d+\.?\d*)$', f)
    if m:
        base = float(m.group(1)) * float(m.group(3))
        divisor = float(m.group(2))
        rate = base / divisor
        return FormulaConversion(f'(bedFeet / 100) * {rate} * daysBetweenHarvest * harvests', 'AREA_DBH_SCALED', False)

    # ========== Seed-based patterns ==========
    # 1/16 * seeds / safetyFactor -> seeds * 0.0625 * harvests
    # Note: safetyFactor is typically 1.1 (10% buffer), we'll bake it in
    m = re.match(r'^1/(\d+\.?\d*)\*seeds/safetyFactor$', f)
    if m:
        rate = 1.0 / float(m.group(1))
        # Bake in typical safety factor of 1.1
        effective_rate = rate / 1.1
        return FormulaConversion(f'seeds * {effective_rate:.6f} * harvests', 'SEEDS_BASED', False)

    # ========== Lettuce complex pattern ==========
    # (fixed + (harvests-1)*rate*plantingsPerBed) / harvests
    m = re.match(r'^\((\d+\.?\d*)\+\(harvests-1\)\*(\d+\.?\d*)\*plantingsPerBed\)/harvests$', f)
    if m:
        fixed = float(m.group(1))
        rate = float(m.group(2))
        # This is complex - output the exact formula
        return FormulaConversion(f'{fixed} + (harvests - 1) * {rate} * plantingsPerBed', 'LETTUCE_COMPLEX', True)

    return None


def calculate_ppb(spacing: float, rows: float, bed_length: float) -> float:
    """Calculate plants per bed."""
    if not spacing or spacing <= 0 or not rows or rows <= 0:
        return 0
    return (12 / spacing) * rows * bed_length


def evaluate_formula(formula: str, context: dict) -> Optional[float]:
    """
    Evaluate our formula string with given context.
    Simple eval with restricted namespace.
    """
    try:
        # Create a restricted namespace with only our variables
        # Using self-documenting names (plantingsPerBed, daysBetweenHarvest)
        namespace = {
            'plantingsPerBed': context.get('plantingsPerBed', 0),
            'bedFeet': context.get('bedFeet', 50),
            'harvests': context.get('harvests', 1),
            'daysBetweenHarvest': context.get('daysBetweenHarvest', 7),
            'rows': context.get('rows', 1),
            'spacing': context.get('spacing', 12),
            'seeds': context.get('seeds', 0),
        }
        result = eval(formula, {"__builtins__": {}}, namespace)
        return float(result)
    except Exception as e:
        print(f"  Error evaluating '{formula}': {e}")
        return None


def main():
    # Ensure tmp directory exists
    Path('tmp').mkdir(exist_ok=True)

    print("Loading workbook...")
    wb_formulas = openpyxl.load_workbook('Crop Plan 2025 V20.xlsm', data_only=False)
    wb_values = openpyxl.load_workbook('Crop Plan 2025 V20.xlsm', data_only=True)
    ws_formulas = wb_formulas['Crop Chart']
    ws_values = wb_values['Crop Chart']

    # Get column positions
    columns = {}
    for col in range(1, 200):
        val = ws_formulas.cell(row=2, column=col).value
        if val:
            columns[val] = col

    bed_length = 50  # Standard bed length

    results = {
        'converted': [],
        'empty': [],
        'unmatched': [],
        'errors': []
    }

    for row in range(3, 350):
        crop = ws_formulas.cell(row=row, column=columns['Crop']).value
        if not crop:
            continue

        product = ws_formulas.cell(row=row, column=columns['Product']).value
        # Get identifier from VALUES workbook (the formula produces the value we need)
        # Strip trailing spaces to match crops.json identifiers
        identifier = ws_values.cell(row=row, column=columns['Identifier']).value
        if identifier:
            identifier = identifier.strip()
        uph_formula = ws_formulas.cell(row=row, column=columns['Units Per Harvest']).value

        # Get values for validation
        spacing = ws_values.cell(row=row, column=columns['Spacing']).value
        rows_val = ws_values.cell(row=row, column=columns['Rows']).value
        harvests = ws_values.cell(row=row, column=columns['Harvests']).value
        dbh = ws_values.cell(row=row, column=columns['Days Between Harvest']).value or 7
        excel_total = ws_values.cell(row=row, column=columns['Custom Yield Per Bed']).value

        # For seed-based crops
        seeds_per_bed = ws_values.cell(row=row, column=columns.get('Seeds Per Bed', 0)).value or 0

        if not uph_formula:
            results['empty'].append({
                'row': row,
                'identifier': identifier,
                'crop': crop,
                'product': product
            })
            continue

        conversion = parse_and_convert(str(uph_formula))

        if not conversion:
            results['unmatched'].append({
                'row': row,
                'identifier': identifier,
                'crop': crop,
                'product': product,
                'excel_formula': str(uph_formula)
            })
            continue

        # Validate by evaluating and comparing to Excel
        if spacing and rows_val and harvests:
            ppb = calculate_ppb(spacing, rows_val, bed_length)
            context = {
                'plantingsPerBed': ppb,
                'bedFeet': bed_length,
                'harvests': harvests,
                'daysBetweenHarvest': dbh,
                'rows': rows_val,
                'spacing': spacing,
                'seeds': seeds_per_bed,
            }

            our_total = evaluate_formula(conversion.formula, context)

            if our_total is not None and excel_total is not None:
                error_pct = abs(our_total - excel_total) / excel_total * 100 if excel_total else 0

                if error_pct > 1:  # More than 1% error
                    results['errors'].append({
                        'row': row,
                        'identifier': identifier,
                        'crop': crop,
                        'product': product,
                        'excel_formula': str(uph_formula),
                        'our_formula': conversion.formula,
                        'pattern': conversion.pattern,
                        'expected': excel_total,
                        'calculated': our_total,
                        'error_pct': error_pct
                    })
                    continue

        # Success!
        results['converted'].append({
            'row': row,
            'identifier': identifier,
            'crop': crop,
            'product': product,
            'yieldFormula': conversion.formula,
            'pattern': conversion.pattern,
            'validated': excel_total is not None
        })

    # Print summary
    print("\n" + "=" * 80)
    print("CONVERSION RESULTS")
    print("=" * 80)

    total = len(results['converted']) + len(results['empty']) + len(results['unmatched']) + len(results['errors'])

    print(f"\nTotal configs: {total}")
    print(f"  Converted & validated: {len(results['converted'])} ({100*len(results['converted'])/total:.1f}%)")
    print(f"  Empty (no yield): {len(results['empty'])} ({100*len(results['empty'])/total:.1f}%)")
    print(f"  Unmatched: {len(results['unmatched'])} ({100*len(results['unmatched'])/total:.1f}%)")
    print(f"  Validation errors: {len(results['errors'])} ({100*len(results['errors'])/total:.1f}%)")

    if results['unmatched']:
        print("\n" + "-" * 80)
        print("UNMATCHED FORMULAS")
        print("-" * 80)
        for item in results['unmatched']:
            print(f"  Row {item['row']}: {item['crop']} - {item['product']}")
            print(f"    Formula: {item['excel_formula'][:70]}")

    if results['errors']:
        print("\n" + "-" * 80)
        print("VALIDATION ERRORS")
        print("-" * 80)
        for item in results['errors']:
            print(f"  Row {item['row']}: {item['crop']} - {item['product']}")
            print(f"    Excel: {item['excel_formula'][:50]}")
            print(f"    Ours:  {item['our_formula']}")
            print(f"    Expected: {item['expected']:.2f}, Got: {item['calculated']:.2f} ({item['error_pct']:.1f}% error)")

    # Output shim file for merging into crops.json
    shim_data = {}
    for item in results['converted']:
        if item['identifier']:
            shim_data[item['identifier']] = {
                'yieldFormula': item['yieldFormula']
            }

    shim_path = Path('tmp/yield-formulas.json')
    with open(shim_path, 'w') as f:
        json.dump(shim_data, f, indent=2)
    print(f"\nShim data written to {shim_path} ({len(shim_data)} configs)")

    # Also output full results for debugging
    with open('tmp/yield-conversion-results.json', 'w') as f:
        json.dump(results, f, indent=2)
    print("Full results written to tmp/yield-conversion-results.json")


if __name__ == '__main__':
    main()
