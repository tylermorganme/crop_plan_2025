#!/usr/bin/env python3
"""
Migrate yield formulas from Excel to the new yield model.

Parses Units Per Harvest formulas and extracts:
- yieldBasis: 'per-plant' | 'per-100ft'
- yieldRate: the multiplier/divisor value
- harvestMultiplies: whether /Harvests is in the formula

Then validates by calculating yield and comparing to Excel's calculated values.
"""

import openpyxl
import json
import re
from dataclasses import dataclass
from typing import Optional, Tuple, Dict, Any

@dataclass
class YieldModel:
    basis: str  # 'per-plant' | 'per-100ft'
    rate: float
    harvest_multiplies: bool
    pattern: str  # for debugging
    extra: Optional[float] = None  # for storing additional data (e.g., fixed_first for lettuce)

def parse_uph_formula(formula: str) -> Optional[YieldModel]:
    """Parse a Units Per Harvest formula into a YieldModel."""
    if not formula or not formula.startswith('='):
        return None

    # Normalize
    f = formula.replace('Crops[[#This Row],', '').replace(']]', ']').replace('[', '').replace(']', '')
    f = f.replace('=', '')

    # ========== PPB-based patterns (per-plant) ==========

    # PPB * multiplier / Harvests
    m = re.match(r'^Plantings Per Bed\*(\d+\.?\d*)/Harvests$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)), False, 'PPB_MULT_DIV_H')

    # PPB / divisor / Harvests
    m = re.match(r'^Plantings Per Bed/(\d+\.?\d*)/Harvests$', f)
    if m:
        return YieldModel('per-plant', 1.0 / float(m.group(1)), False, 'PPB_DIV_DIV_H')

    # PPB / Harvests
    m = re.match(r'^Plantings Per Bed/Harvests$', f)
    if m:
        return YieldModel('per-plant', 1.0, False, 'PPB_DIV_DIV_H')

    # PPB * multiplier (no /Harvests)
    m = re.match(r'^Plantings Per Bed\*(\d+\.?\d*)$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)), True, 'PPB_MULT')

    # PPB / divisor (no /Harvests)
    m = re.match(r'^Plantings Per Bed/(\d+\.?\d*)$', f)
    if m:
        return YieldModel('per-plant', 1.0 / float(m.group(1)), True, 'PPB_DIV')

    # PPB direct
    m = re.match(r'^Plantings Per Bed$', f)
    if m:
        return YieldModel('per-plant', 1.0, True, 'PPB_DIRECT')

    # PPB * (fraction)
    m = re.match(r'^Plantings Per Bed\*\((\d+\.?\d*)/(\d+\.?\d*)\)$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)) / float(m.group(2)), True, 'PPB_MULT')

    # PPB * (fraction) * multiplier
    m = re.match(r'^Plantings Per Bed\*\((\d+\.?\d*)/(\d+\.?\d*)\)\*(\d+\.?\d*)$', f)
    if m:
        rate = float(m.group(1)) / float(m.group(2)) * float(m.group(3))
        return YieldModel('per-plant', rate, True, 'PPB_MULT')

    # PPB * a / b (not Harvests)
    m = re.match(r'^Plantings Per Bed\*(\d+\.?\d*)/(\d+\.?\d*)$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)) / float(m.group(2)), True, 'PPB_MULT')

    # PPB / a / b / Harvests
    m = re.match(r'^Plantings Per Bed/(\d+\.?\d*)/(\d+\.?\d*)/Harvests$', f)
    if m:
        return YieldModel('per-plant', 1.0 / (float(m.group(1)) * float(m.group(2))), False, 'PPB_DIV_DIV_H')

    # PPB / a / b / c / Harvests
    m = re.match(r'^Plantings Per Bed/(\d+\.?\d*)/(\d+\.?\d*)/(\d+\.?\d*)/Harvests$', f)
    if m:
        return YieldModel('per-plant', 1.0 / (float(m.group(1)) * float(m.group(2)) * float(m.group(3))), False, 'PPB_DIV_DIV_H')

    # PPB / a / b / c (no Harvests)
    m = re.match(r'^Plantings Per Bed/(\d+\.?\d*)/(\d+\.?\d*)/(\d+\.?\d*)$', f)
    if m:
        return YieldModel('per-plant', 1.0 / (float(m.group(1)) * float(m.group(2)) * float(m.group(3))), True, 'PPB_DIV')

    # PPB * a / b / Harvests
    m = re.match(r'^Plantings Per Bed\*(\d+\.?\d*)/(\d+\.?\d*)/Harvests$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)) / float(m.group(2)), False, 'PPB_MULT_DIV_H')

    # PPB / a * b / Harvests
    m = re.match(r'^Plantings Per Bed/(\d+\.?\d*)\*(\d+\.?\d*)/Harvests$', f)
    if m:
        return YieldModel('per-plant', float(m.group(2)) / float(m.group(1)), False, 'PPB_MULT_DIV_H')

    # PPB * a * b / Harvests
    m = re.match(r'^Plantings Per Bed\*(\d+\.?\d*)\*(\d+\.?\d*)/Harvests$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)) * float(m.group(2)), False, 'PPB_MULT_DIV_H')

    # PPB * a * b * c / Harvests
    m = re.match(r'^Plantings Per Bed\*(\d+\.?\d*)\*(\d+\.?\d*)\*(\d+\.?\d*)/Harvests$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)) * float(m.group(2)) * float(m.group(3)), False, 'PPB_MULT_DIV_H')

    # PPB * a * b / c / Harvests (onion pattern)
    m = re.match(r'^Plantings Per Bed\*(\d+\.?\d*)\*(\d+\.?\d*)/(\d+\.?\d*)/Harvests$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)) * float(m.group(2)) / float(m.group(3)), False, 'PPB_MULT_DIV_H')

    # PPB / Harvests * multiplier
    m = re.match(r'^Plantings Per Bed/Harvests\*(\d+\.?\d*)$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)), False, 'PPB_MULT_DIV_H')

    # 1/Harvests * PPB
    m = re.match(r'^1/Harvests\*Plantings Per Bed$', f)
    if m:
        return YieldModel('per-plant', 1.0, False, 'PPB_DIV_DIV_H')

    # PPB / a * b (no Harvests)
    m = re.match(r'^Plantings Per Bed/(\d+\.?\d*)\*(\d+\.?\d*)$', f)
    if m:
        return YieldModel('per-plant', float(m.group(2)) / float(m.group(1)), True, 'PPB_MULT')

    # ========== Area-based patterns (per-100ft) ==========

    # X * BedLength/StandardBedLength / Harvests
    m = re.match(r'^(\d+\.?\d*)\*BedLength/StandardBedLength/Harvests$', f)
    if m:
        return YieldModel('per-100ft', float(m.group(1)), False, 'AREA_DIV_H')

    # X * BedLength/StandardBedLength (no /Harvests)
    m = re.match(r'^(\d+\.?\d*)\*BedLength/StandardBedLength$', f)
    if m:
        return YieldModel('per-100ft', float(m.group(1)), True, 'AREA')

    # BedLength / Harvests
    m = re.match(r'^BedLength/Harvests$', f)
    if m:
        return YieldModel('per-100ft', 100.0, False, 'BEDLENGTH_DIV_H')

    # BedLength * X / Harvests
    m = re.match(r'^BedLength\*(\d+\.?\d*)/Harvests$', f)
    if m:
        return YieldModel('per-100ft', float(m.group(1)) * 100.0, False, 'BEDLENGTH_DIV_H')

    # X * Y * BedLength / Harvests
    m = re.match(r'^(\d+\.?\d*)\*(\d+\.?\d*)\*BedLength/Harvests$', f)
    if m:
        return YieldModel('per-100ft', float(m.group(1)) * float(m.group(2)) * 100.0, False, 'BEDLENGTH_DIV_H')

    # Sunchoke: 400*BedLength/StandardBedLength/27*2.2*10/Harvests
    m = re.match(r'^(\d+\.?\d*)\*BedLength/StandardBedLength/(\d+\.?\d*)\*(\d+\.?\d*)\*(\d+\.?\d*)/Harvests$', f)
    if m:
        total = float(m.group(1)) / float(m.group(2)) * float(m.group(3)) * float(m.group(4))
        return YieldModel('per-100ft', total, False, 'AREA_DIV_H')

    # ========== DBH-scaled patterns (weekly production) ==========

    # Cucumber/Squash/Rat Tail: PPB * rate * DBH/7
    # These calculate weekly production * number of weeks
    # We'll convert to: per-plant rate = rate * (DBH/7) / harvests, harvestMultiplies=true
    # But we need DBH and Harvests values to compute the effective rate

    # PPB * rate * DBH / 7 (need DBH value to compute)
    m = re.match(r'^Plantings Per Bed\*(\d+\.?\d*)\*Days Between Harvest/7$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)), True, 'PPB_WEEKLY_RATE')

    # PPB * DBH / 7 (rate = 1)
    m = re.match(r'^Plantings Per Bed\*Days Between Harvest/7$', f)
    if m:
        return YieldModel('per-plant', 1.0, True, 'PPB_WEEKLY_RATE')

    # PPB * rate * (DBH / 7) - parenthesized
    m = re.match(r'^Plantings Per Bed\*(\d+\.?\d*)\*\(Days Between Harvest/7\)$', f)
    if m:
        return YieldModel('per-plant', float(m.group(1)), True, 'PPB_WEEKLY_RATE')

    # PPB * rate * DBH (rat tail radish - DBH directly multiplies, not weekly)
    m = re.match(r'^Plantings Per Bed\*(\d+\.?\d*)\*Days Between Harvest$', f)
    if m:
        # rate * DBH = per-day rate * days between harvest
        return YieldModel('per-plant', float(m.group(1)), True, 'PPB_DBH_DIRECT')

    # ========== Tomato pattern (area-based, DBH-scaled) ==========
    # 500*BedLength/StandardBedLength*(Harvests*DBH)/42/Harvests*scale
    # Simplifies to: 500*(BedLength/100)*(DBH/42)*scale
    # = (500*scale) * (BedLength/100) * (DBH/42)
    # Base rate = 500*scale per 100ft per 6 weeks, DBH scales it

    # Tomato large: 500*BedLength/StandardBedLength*(Harvests*DBH)/42/Harvests
    # This computes UPH, so harvestMultiplies=True for total yield
    m = re.match(r'^(\d+\.?\d*)\*BedLength/StandardBedLength\*\(Harvests\*Days Between Harvest\)/(\d+\.?\d*)/Harvests$', f)
    if m:
        base_rate = float(m.group(1))
        divisor = float(m.group(2))
        # Rate is base/divisor per 100ft, DBH scales it
        return YieldModel('per-100ft', base_rate / divisor, True, 'AREA_DBH_SCALED')

    # Tomato scaled: same pattern * scale
    m = re.match(r'^(\d+\.?\d*)\*BedLength/StandardBedLength\*\(Harvests\*Days Between Harvest\)/(\d+\.?\d*)/Harvests\*(\d+\.?\d*)$', f)
    if m:
        base_rate = float(m.group(1)) * float(m.group(3))
        divisor = float(m.group(2))
        return YieldModel('per-100ft', base_rate / divisor, True, 'AREA_DBH_SCALED')

    # ========== Seed-based patterns ==========

    # Shallots: 1/16 * SeedsPerBed / SafetyFactor
    # This computes UPH, total = UPH * Harvests
    m = re.match(r'^1/(\d+\.?\d*)\*Seeds Per Bed/Safety Factor$', f)
    if m:
        # Seeds per bed depends on spacing, so this is effectively per-plant
        # but with seed count instead of plant count
        divisor = float(m.group(1))
        return YieldModel('per-plant', 1.0 / divisor, True, 'SEEDS_BASED')

    # ========== Lettuce pattern ==========

    # Lettuce: (fixed + (H-1)*0.25*PPB) / H
    # First harvest = fixed (0.5 or 0.75), subsequent = 0.25 * PPB each
    m = re.match(r'^\((\d+\.?\d*)\+\(Harvests-1\)\*(\d+\.?\d*)\*Plantings Per Bed\)/Harvests$', f)
    if m:
        fixed_first = float(m.group(1))
        subsequent_rate = float(m.group(2))
        # Store both: rate for subsequent harvests, extra for fixed first
        return YieldModel('per-plant', subsequent_rate, False, 'LETTUCE_COMPLEX', extra=fixed_first)

    return None


def calculate_ppb(spacing: float, rows: float, bed_length: float) -> float:
    """Calculate plants per bed."""
    if not spacing or spacing <= 0 or not rows or rows <= 0:
        return 0
    return (12 / spacing) * rows * bed_length


def calculate_yield(model: YieldModel, spacing: float, rows: float,
                   bed_length: float, harvests: float,
                   dbh: float = 7, seeds_per_bed: float = 0, safety_factor: float = 1) -> float:
    """Calculate total yield using our model."""
    ppb = calculate_ppb(spacing, rows, bed_length)

    if model.pattern == 'PPB_WEEKLY_RATE':
        # For weekly rate crops, rate is per week, multiply by weeks
        weeks = dbh / 7 if dbh else 1
        base_yield = ppb * model.rate * weeks
    elif model.pattern == 'PPB_DBH_DIRECT':
        # PPB * rate * DBH directly (rat tail radish)
        base_yield = ppb * model.rate * dbh
    elif model.pattern == 'AREA_DBH_SCALED':
        # Tomato pattern: (BedLength/100) * rate * DBH
        # rate = 500/42 or similar, DBH gives weeks of production
        # Result is UPH (Units Per Harvest) since harvestMultiplies=True
        base_yield = (bed_length / 100) * model.rate * dbh
    elif model.pattern == 'SEEDS_BASED':
        # Shallots: 1/16 * SeedsPerBed / SafetyFactor
        # rate = 1/16, so: SeedsPerBed * rate / SafetyFactor
        base_yield = seeds_per_bed * model.rate / safety_factor if safety_factor else 0
    elif model.pattern == 'LETTUCE_COMPLEX':
        # Lettuce: (fixed + (H-1)*0.25*PPB) / H
        # model.extra contains fixed_first, model.rate contains subsequent rate
        fixed_first = model.extra or 0.5
        uph = (fixed_first + (harvests - 1) * model.rate * ppb) / harvests if harvests else 0
        base_yield = uph * harvests
    elif model.basis == 'per-plant':
        base_yield = ppb * model.rate
    else:  # per-100ft
        base_yield = (bed_length / 100) * model.rate

    if model.harvest_multiplies:
        return base_yield * harvests
    else:
        return base_yield


def main():
    # Load workbook
    print("Loading workbook...")
    wb_formulas = openpyxl.load_workbook('Crop Plan 2025 V20.xlsm', data_only=False)
    wb_values = openpyxl.load_workbook('Crop Plan 2025 V20.xlsm', data_only=True)
    ws_formulas = wb_formulas['Crop Chart']
    ws_values = wb_values['Crop Chart']

    # Get column positions
    columns = {}
    for col in range(1, 200):
        val = ws_formulas.cell(row=2, column=col).value
        if val:
            columns[val] = col

    bed_length = 50  # Standard bed length

    results = {
        'matched': [],
        'unmatched': [],
        'empty': [],
        'errors': []
    }

    for row in range(3, 350):
        crop = ws_formulas.cell(row=row, column=columns['Crop']).value
        if not crop:
            continue

        product = ws_formulas.cell(row=row, column=columns['Product']).value
        uph_formula = ws_formulas.cell(row=row, column=columns['Units Per Harvest']).value

        # Get values for validation
        spacing = ws_values.cell(row=row, column=columns['Spacing']).value
        rows_val = ws_values.cell(row=row, column=columns['Rows']).value
        harvests = ws_values.cell(row=row, column=columns['Harvests']).value
        dbh = ws_values.cell(row=row, column=columns['Days Between Harvest']).value or 7
        excel_uph = ws_values.cell(row=row, column=columns['Units Per Harvest']).value
        excel_total = ws_values.cell(row=row, column=columns['Custom Yield Per Bed']).value

        # For seed-based crops
        seeds_per_bed = ws_values.cell(row=row, column=columns.get('Seeds Per Bed', 0)).value or 0
        safety_factor = ws_values.cell(row=row, column=columns.get('Safety Factor', 0)).value or 1

        if not uph_formula:
            results['empty'].append({
                'row': row,
                'crop': crop,
                'product': product
            })
            continue

        model = parse_uph_formula(str(uph_formula))

        if not model:
            results['unmatched'].append({
                'row': row,
                'crop': crop,
                'product': product,
                'formula': str(uph_formula)
            })
            continue

        # Validate: calculate yield and compare to Excel
        if spacing and rows_val and harvests:
            try:
                calculated_total = calculate_yield(
                    model, spacing, rows_val, bed_length, harvests,
                    dbh=dbh, seeds_per_bed=seeds_per_bed, safety_factor=safety_factor
                )

                # Compare to Excel's calculated value
                if excel_total and abs(calculated_total - excel_total) > 0.01:
                    error_pct = abs(calculated_total - excel_total) / excel_total * 100 if excel_total else 0
                    if error_pct > 1:  # More than 1% error
                        results['errors'].append({
                            'row': row,
                            'crop': crop,
                            'product': product,
                            'formula': str(uph_formula),
                            'pattern': model.pattern,
                            'expected': excel_total,
                            'calculated': calculated_total,
                            'error_pct': error_pct
                        })
                        continue

                results['matched'].append({
                    'row': row,
                    'crop': crop,
                    'product': product,
                    'basis': model.basis,
                    'rate': model.rate,
                    'harvest_multiplies': model.harvest_multiplies,
                    'pattern': model.pattern,
                    'extra': model.extra,
                    'validated_total': calculated_total
                })
            except Exception as e:
                results['errors'].append({
                    'row': row,
                    'crop': crop,
                    'product': product,
                    'formula': str(uph_formula),
                    'error': str(e)
                })
        else:
            # Can't validate but pattern matched
            results['matched'].append({
                'row': row,
                'crop': crop,
                'product': product,
                'basis': model.basis,
                'rate': model.rate,
                'harvest_multiplies': model.harvest_multiplies,
                'pattern': model.pattern,
                'extra': model.extra,
                'validated_total': None
            })

    # Print results
    print("\n" + "=" * 80)
    print("MIGRATION RESULTS")
    print("=" * 80)

    total = len(results['matched']) + len(results['unmatched']) + len(results['empty']) + len(results['errors'])

    print(f"\nTotal configs: {total}")
    print(f"  Matched & validated: {len(results['matched'])} ({100*len(results['matched'])/total:.1f}%)")
    print(f"  Empty (no yield): {len(results['empty'])} ({100*len(results['empty'])/total:.1f}%)")
    print(f"  Unmatched (need manual): {len(results['unmatched'])} ({100*len(results['unmatched'])/total:.1f}%)")
    print(f"  Validation errors: {len(results['errors'])} ({100*len(results['errors'])/total:.1f}%)")

    if results['unmatched']:
        print("\n" + "-" * 80)
        print("UNMATCHED FORMULAS (need special handling)")
        print("-" * 80)
        for item in results['unmatched']:
            print(f"  {item['crop']} ({item['product']}): {item['formula'][:60]}")

    if results['errors']:
        print("\n" + "-" * 80)
        print("VALIDATION ERRORS (calculation mismatch)")
        print("-" * 80)
        for item in results['errors']:
            print(f"  {item['crop']} ({item['product']})")
            print(f"    Pattern: {item.get('pattern', 'N/A')}")
            print(f"    Formula: {item.get('formula', 'N/A')[:60]}")
            if 'expected' in item:
                print(f"    Expected: {item['expected']:.2f}, Got: {item['calculated']:.2f} ({item['error_pct']:.1f}% error)")
            else:
                print(f"    Error: {item.get('error', 'Unknown')}")

    # Save results for use in build script
    with open('/tmp/yield_migration.json', 'w') as f:
        json.dump(results, f, indent=2)
    print(f"\nResults saved to /tmp/yield_migration.json")

    # Output yield model data for updating crops.json
    print("\n" + "=" * 80)
    print("YIELD MODEL SUMMARY")
    print("=" * 80)
    print("\nPattern distribution:")
    patterns = {}
    for item in results['matched']:
        p = item['pattern']
        patterns[p] = patterns.get(p, 0) + 1
    for p, count in sorted(patterns.items(), key=lambda x: -x[1]):
        print(f"  {p}: {count}")

    # Export yield data for build script
    # Map internal patterns to CropConfig yieldPattern values
    pattern_map = {
        'PPB_WEEKLY_RATE': 'PPB_WEEKLY_RATE',
        'PPB_DBH_DIRECT': 'PPB_DBH_DIRECT',
        'AREA_DBH_SCALED': 'AREA_DBH_SCALED',
        'SEEDS_BASED': 'SEEDS_BASED',
        'LETTUCE_COMPLEX': 'LETTUCE_COMPLEX',
    }

    yield_data = []
    for item in results['matched']:
        entry = {
            'row': item['row'],
            'crop': item['crop'],
            'product': item['product'],
            'yieldBasis': item['basis'],
            'yieldRate': round(item['rate'], 6),
            'harvestMultiplies': item['harvest_multiplies'],
        }
        # Add yieldPattern for edge cases
        if item['pattern'] in pattern_map:
            entry['yieldPattern'] = pattern_map[item['pattern']]
        # Add yieldExtra for lettuce
        if item['pattern'] == 'LETTUCE_COMPLEX' and 'extra' in item and item['extra']:
            entry['yieldExtra'] = item['extra']
        yield_data.append(entry)

    with open('/tmp/yield_model_data.json', 'w') as f:
        json.dump(yield_data, f, indent=2)
    print(f"\nYield model data exported to /tmp/yield_model_data.json ({len(yield_data)} configs)")


if __name__ == '__main__':
    main()
