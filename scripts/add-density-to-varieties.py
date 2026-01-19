#!/usr/bin/env python3
"""
Add density data from Excel Varieties sheet to varieties-template.json.

This script:
1. Reads density data from the Excel Varieties sheet
2. Matches varieties by crop|name|supplier key
3. Updates varieties-template.json with density and densityUnit fields
"""

import openpyxl
import json
from datetime import datetime

# Unit mapping from Excel to our DensityUnit type
UNIT_MAP = {
    'g': 'g',
    'ozm': 'oz',
    'lbm': 'lb',
    'Count Only': 'ct',
    'By Count': 'ct',
}

def extract_density_from_excel(excel_path: str) -> dict:
    """Extract density data from Excel Varieties sheet."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Varieties']

    density_data = {}

    for row in range(3, 800):
        crop = ws.cell(row=row, column=3).value
        name = ws.cell(row=row, column=5).value
        supplier = ws.cell(row=row, column=7).value
        density = ws.cell(row=row, column=17).value
        density_per = ws.cell(row=row, column=18).value

        if not crop or not name or not supplier:
            continue

        # Skip if no density data
        if not density or density == '#N/A' or density_per == '#N/A':
            continue

        # Create key matching variety ID format
        key = f'{crop.lower().strip()}|{name.lower().strip()}|{supplier.lower().strip()}'

        unit = UNIT_MAP.get(density_per)
        if not unit:
            continue

        # Handle count-only varieties
        if unit == 'ct' and (not isinstance(density, (int, float)) or density == 0):
            density_data[key] = {'density': None, 'densityUnit': 'ct'}
        else:
            density_data[key] = {'density': density, 'densityUnit': unit}

    return density_data


def update_varieties_json(varieties_path: str, density_data: dict) -> dict:
    """Update varieties.json with density data."""
    with open(varieties_path) as f:
        data = json.load(f)

    varieties = data['varieties']
    matched = 0
    unmatched = 0

    for variety in varieties:
        key = f"{variety['crop'].lower().strip()}|{variety['name'].lower().strip()}|{variety['supplier'].lower().strip()}"

        if key in density_data:
            density_info = density_data[key]
            if density_info['density'] is not None:
                variety['density'] = density_info['density']
            variety['densityUnit'] = density_info['densityUnit']
            matched += 1
        else:
            unmatched += 1

    # Update metadata
    data['_generated'] = datetime.utcnow().isoformat() + 'Z'

    return {
        'data': data,
        'matched': matched,
        'unmatched': unmatched,
        'total_density_records': len(density_data),
    }


def main():
    excel_path = 'Crop Plan 2025 V20.xlsm'
    varieties_path = 'src/data/varieties-template.json'

    print(f'Extracting density data from {excel_path}...')
    density_data = extract_density_from_excel(excel_path)
    print(f'Found {len(density_data)} density records')

    print(f'\nUpdating {varieties_path}...')
    result = update_varieties_json(varieties_path, density_data)

    print(f'Matched: {result["matched"]}')
    print(f'Unmatched: {result["unmatched"]}')

    # Write updated data
    with open(varieties_path, 'w') as f:
        json.dump(result['data'], f, indent=2)

    print(f'\nUpdated {varieties_path}')

    # Show samples of updated varieties
    print('\nSample updated varieties:')
    count = 0
    for v in result['data']['varieties']:
        if 'density' in v or 'densityUnit' in v:
            print(f"  {v['crop']} - {v['name']} ({v['supplier']}): {v.get('density')} {v.get('densityUnit')}")
            count += 1
            if count >= 5:
                break


if __name__ == '__main__':
    main()
