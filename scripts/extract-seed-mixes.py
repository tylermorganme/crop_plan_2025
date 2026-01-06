#!/usr/bin/env python3
"""
Extract seed mix data from the Seed Mixes sheet in the Excel workbook.
Creates a JSON file for the seed mix import pipeline.

Usage:
    python scripts/extract-seed-mixes.py
"""

import openpyxl
import json
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

# Custom JSON encoder for datetime objects
class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)

def fmt_value(val):
    """Format a cell value for JSON output"""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip() if val.strip() else None
    return val

# Load the workbook
print("Loading workbook...")
wb = openpyxl.load_workbook("Crop Plan 2025 V20.xlsm", data_only=True)

# Check available sheets
print(f"Available sheets: {wb.sheetnames}")

# Read the Seed Mixes sheet
if "Seed Mixes" not in wb.sheetnames:
    print("ERROR: 'Seed Mixes' sheet not found!")
    exit(1)

ws = wb["Seed Mixes"]

# Get headers from row 1
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value
    headers.append(header)

print(f"Found {len(headers)} columns: {headers}")

# Expected columns: Percent, Name, Crop, Variety, Company, Label

# Extract mix components starting from row 2
# Each row is a component of a mix (mix name can repeat across multiple rows)
raw_components = []
for row in range(2, ws.max_row + 1):
    row_data = {}
    for col, header in enumerate(headers, start=1):
        if header:
            value = ws.cell(row=row, column=col).value
            row_data[header] = fmt_value(value)

    # Skip empty rows
    mix_name = row_data.get("Name") or row_data.get("Mix Name")
    if not mix_name:
        continue

    component = {
        "mixName": mix_name,
        "crop": row_data.get("Crop") or "",
        "variety": row_data.get("Variety") or "",
        "supplier": row_data.get("Company") or row_data.get("Supplier") or "",
        "percent": row_data.get("Percent") or row_data.get("Percentage") or 0,
        "label": row_data.get("Label") or "",
    }

    raw_components.append(component)

print(f"Extracted {len(raw_components)} mix components")

# Group components by mix name to create seed mixes
mixes_by_name = defaultdict(list)
for comp in raw_components:
    mix_name = comp["mixName"]
    mixes_by_name[mix_name].append(comp)

# Build seed mix objects
seed_mixes = []
for mix_name, components in mixes_by_name.items():
    # Get crop from first component (should all be same crop)
    crop = components[0].get("crop", "")

    mix = {
        "name": mix_name,
        "crop": crop,
        "components": [
            {
                "variety": c["variety"],
                "supplier": c["supplier"],
                "percent": c["percent"],
                "label": c.get("label", ""),
            }
            for c in components
        ]
    }

    seed_mixes.append(mix)

print(f"Created {len(seed_mixes)} seed mixes")

# Ensure output directory exists
output_dir = Path("tmp")
output_dir.mkdir(exist_ok=True)

# Save to JSON
output_path = output_dir / "seed_mixes_from_excel.json"
with open(output_path, "w") as f:
    json.dump({"seedMixes": seed_mixes}, f, indent=2, cls=DateTimeEncoder)

print(f"Saved to {output_path}")

# Show sample and stats
print("\n--- Statistics ---")
crops = set(m.get("crop", "") for m in seed_mixes if m.get("crop"))
avg_components = sum(len(m["components"]) for m in seed_mixes) / len(seed_mixes) if seed_mixes else 0

print(f"Unique crops: {len(crops)}")
print(f"Avg components per mix: {avg_components:.1f}")

print("\nSample seed mix:")
if seed_mixes:
    print(json.dumps(seed_mixes[0], indent=2))
