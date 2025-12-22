"""
Trace the formula dependency DAG in the Bed Plan sheet.
Starting from "End of Harvest" (col 36), trace back through all dependencies
to build a complete picture of the calculation chain.
"""

import openpyxl
from openpyxl.formula import Tokenizer
import re
from collections import defaultdict

# Load workbook WITH formulas (not data_only)
wb = openpyxl.load_workbook("Crop Plan 2025 V20.xlsm", data_only=False)
ws = wb["Bed Plan"]

# Get headers for reference
headers = {}
for col in range(1, 160):
    val = ws.cell(row=5, column=col).value
    if val:
        headers[col] = val

def col_letter_to_num(col_str):
    """Convert column letter(s) to number (A=1, B=2, ..., AA=27)"""
    result = 0
    for char in col_str.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

def parse_cell_ref(ref):
    """Parse a cell reference like 'A6', '$B$6', 'AJ6' into (col_num, row_num)"""
    # Remove sheet references
    if '!' in ref:
        ref = ref.split('!')[-1]

    # Remove $ signs
    ref = ref.replace('$', '')

    # Match column letters and row number
    match = re.match(r'^([A-Za-z]+)(\d+)$', ref)
    if match:
        col_str, row_str = match.groups()
        return col_letter_to_num(col_str), int(row_str)
    return None, None

def get_cell_refs_from_formula(formula):
    """Extract all cell references from a formula using Tokenizer"""
    if not formula or not str(formula).startswith('='):
        return []

    try:
        tok = Tokenizer(formula)
        refs = []
        for token in tok.items:
            if token.type == 'OPERAND' and token.subtype == 'RANGE':
                # Could be a single cell or a range
                val = token.value
                if ':' in val:
                    # Range like A1:A10 - just take the start for now
                    val = val.split(':')[0]
                refs.append(val)
        return refs
    except Exception as e:
        print(f"  Error parsing formula: {e}")
        return []

def trace_dependencies(start_col, row=6, max_depth=20):
    """
    Recursively trace formula dependencies starting from a column.
    Returns a DAG structure.
    """
    visited = set()
    dag = defaultdict(list)  # col -> list of dependent cols

    def trace(col, depth=0):
        if depth > max_depth:
            return
        if col in visited:
            return
        visited.add(col)

        cell = ws.cell(row=row, column=col)
        formula = cell.value

        if not formula or not str(formula).startswith('='):
            return

        refs = get_cell_refs_from_formula(str(formula))

        for ref in refs:
            ref_col, ref_row = parse_cell_ref(ref)
            if ref_col and ref_row == row:  # Same row reference
                dag[col].append(ref_col)
                trace(ref_col, depth + 1)

    trace(start_col)
    return dag

# Trace from End of Harvest (col 36) backwards
print("=" * 80)
print("FORMULA DEPENDENCY DAG - Bed Plan Sheet")
print("Tracing from 'End of Harvest' (col 36) back to inputs")
print("=" * 80)

# First, let's see the actual formulas for the key timing columns
print("\n=== KEY FORMULAS (Row 6) ===\n")

key_cols = [
    36,  # End of Harvest (FINAL OUTPUT)
    35,  # Failed
    33,  # Additional Days of Harvest
    32,  # Actual End of Harvest
    31,  # Expected End of Harvest
    30,  # Beginning of Harvest
    29,  # Actual Beginning of Harvest
    28,  # Expected Beginning of Harvest
    27,  # Additional Days In Field
    26,  # TP or DS Date
    25,  # In Ground Days Late
    24,  # Actual TP or DS Date
    23,  # Planned TP or DS Date
    22,  # Follow Offset
    21,  # Follows Crop
    20,  # Fixed Field Start Date
    19,  # Greenhouse Start Date
    18,  # Actual Greenhouse Date
    17,  # Planned Greenhouse Start Date
    16,  # Start Date
]

for col in key_cols:
    cell = ws.cell(row=6, column=col)
    header = headers.get(col, f"Col {col}")
    formula = cell.value
    print(f"Col {col:2d} | {header}")
    print(f"        Formula: {formula}")
    print()

# Now trace the DAG
print("\n=== DEPENDENCY DAG ===\n")
dag = trace_dependencies(36, row=6)

# Print DAG in readable format
def print_dag(dag, start_col, headers, indent=0):
    """Print DAG as a tree"""
    header = headers.get(start_col, f"Col {start_col}")
    prefix = "  " * indent + ("└─ " if indent > 0 else "")
    print(f"{prefix}[{start_col}] {header}")

    for dep_col in sorted(dag.get(start_col, [])):
        print_dag(dag, dep_col, headers, indent + 1)

print("Starting from End of Harvest:\n")
print_dag(dag, 36, headers)

# Also create a simplified linear flow
print("\n\n=== CALCULATION FLOW (Simplified) ===\n")

# Manually trace the logical flow based on column names
flow = """
INPUTS (User-specified or from Crop Database):
  [16] Start Date ← User sets the target seeding date
  [57] DTM (Days to Maturity) ← From crop database
  [58] Additional Days in Cells ← User adjustment
  [59] Harvest Window ← From crop database
  [60] DS/TP ← Direct Seed or Transplant
  [62] Days in Cells ← From crop database

GREENHOUSE PHASE:
  [17] Planned Greenhouse Start Date = Start Date (if transplant)
  [18] Actual Greenhouse Date ← User override
  [19] Greenhouse Start Date = COALESCE(Actual, Planned)

FIELD PHASE:
  [20] Fixed Field Start Date ← User override for field date
  [21] Follows Crop ← Another crop this one follows
  [22] Follow Offset ← Days after followed crop
  [23] Planned TP or DS Date = Start Date + Days in Cells (+ Additional)
  [24] Actual TP or DS Date ← User override
  [25] In Ground Days Late = Actual - Planned (deviation tracking)
  [26] TP or DS Date = COALESCE(Fixed, Actual, Planned, Follow logic)
  [27] Additional Days In Field ← User adjustment

HARVEST PHASE:
  [28] Expected Beginning of Harvest = TP/DS Date + DTM
  [29] Actual Beginning of Harvest ← User override
  [30] Beginning of Harvest = COALESCE(Actual, Expected)
  [31] Expected End of Harvest = Beginning + Harvest Window
  [32] Actual End of Harvest ← User override
  [33] Additional Days of Harvest ← User adjustment
  [35] Failed ← Mark if crop failed
  [36] End of Harvest = COALESCE(Actual, Expected + Additional) or blank if Failed
"""
print(flow)

# Now let's look at one specific formula in detail
print("\n=== DETAILED FORMULA ANALYSIS ===\n")

# End of Harvest formula
eoh_formula = ws.cell(row=6, column=36).value
print(f"End of Harvest Formula:\n{eoh_formula}\n")

# Expected End of Harvest formula
eeoh_formula = ws.cell(row=6, column=31).value
print(f"Expected End of Harvest Formula:\n{eeoh_formula}\n")

# Beginning of Harvest formula
boh_formula = ws.cell(row=6, column=30).value
print(f"Beginning of Harvest Formula:\n{boh_formula}\n")

# TP or DS Date formula
tpds_formula = ws.cell(row=6, column=26).value
print(f"TP or DS Date Formula:\n{tpds_formula}\n")
