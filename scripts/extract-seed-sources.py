#!/usr/bin/env python3
"""
Extract seed source assignments from Excel Bed Plan sheet.

Reads: Crop Plan 2025 V20.xlsm (Bed Plan sheet)
Writes: tmp/seed_sources_from_excel.json

Maps planting identifiers to their seed source (variety or mix).
"""

import json
import openpyxl
from pathlib import Path

# Paths
EXCEL_PATH = Path('Crop Plan 2025 V20.xlsm')
OUTPUT_PATH = Path('tmp/seed_sources_from_excel.json')

def main():
    print(f'Reading {EXCEL_PATH}...')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Bed Plan']

    # Headers are in row 5
    header_row = 5
    headers = {}
    for col in range(1, 50):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[val] = col

    print(f'Found headers: {list(headers.keys())[:15]}...')

    # Required columns
    id_col = headers.get('Identifier')
    variety_col = headers.get('Variety')
    company_col = headers.get('Company')

    if not id_col or not variety_col:
        print('ERROR: Could not find required columns')
        print(f'  Identifier col: {id_col}')
        print(f'  Variety col: {variety_col}')
        return

    print(f'Using columns: Identifier={id_col}, Variety={variety_col}, Company={company_col}')

    # Extract seed sources
    seed_sources = {}
    row = header_row + 1
    empty_streak = 0

    while empty_streak < 10:
        identifier = ws.cell(row=row, column=id_col).value
        variety = ws.cell(row=row, column=variety_col).value
        company = ws.cell(row=row, column=company_col).value if company_col else None

        if not identifier:
            empty_streak += 1
            row += 1
            continue

        empty_streak = 0

        # Clean up values
        variety = str(variety).strip() if variety else None
        company = str(company).strip() if company else None

        # Skip blank/invalid entries
        if company in ('(blank)', 'None', ''):
            company = None
        if variety in ('(blank)', 'None', ''):
            variety = None

        if variety:
            # Determine if it's a mix or variety
            # It's a mix if the name ends in "Mix" OR if there's no supplier
            is_mix = variety.lower().endswith(' mix') or not company

            seed_sources[identifier] = {
                'variety': variety,
                'supplier': company,
                'isMix': is_mix,
            }

        row += 1

    print(f'Extracted {len(seed_sources)} seed source assignments')

    # Stats
    mixes = sum(1 for s in seed_sources.values() if s['isMix'])
    varieties = len(seed_sources) - mixes
    print(f'  Mixes: {mixes}')
    print(f'  Varieties: {varieties}')

    # Write output
    OUTPUT_PATH.parent.mkdir(exist_ok=True)
    with open(OUTPUT_PATH, 'w') as f:
        json.dump({
            '_source': str(EXCEL_PATH),
            '_sheet': 'Bed Plan',
            'seedSources': seed_sources,
        }, f, indent=2)

    print(f'Wrote {OUTPUT_PATH}')

    # Show some examples
    print('\nSample entries:')
    for i, (id, source) in enumerate(list(seed_sources.items())[:15]):
        type_str = 'MIX' if source['isMix'] else 'VAR'
        supplier_str = source['supplier'] or '-'
        print(f'  {id}: [{type_str}] {source["variety"]} ({supplier_str})')

if __name__ == '__main__':
    main()
