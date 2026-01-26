#!/usr/bin/env python3
"""
Extract crop data from the Crop Chart sheet in the Excel workbook.
Creates a JSON file with crop data including cell colors.

Colors are extracted from the Crop column (column J) cell formatting:
- bgColor: background fill color
- textColor: font color
"""

import openpyxl
from openpyxl.styles import PatternFill
import json
import hashlib
from datetime import datetime, date

# Custom JSON encoder for datetime objects
class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)

def fmt_value(val):
    """Format a cell value for JSON output"""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    return val

def generate_id(identifier):
    """Generate a stable ID from the identifier string"""
    h = hashlib.md5(identifier.encode()).hexdigest()[:8]
    return f"crop_{h}"

def get_theme_color(theme_id):
    """Get RGB hex from Excel theme color index."""
    # Standard Excel Office theme colors
    theme_colors = {
        0: 'ffffff',  # lt1 (background)
        1: '000000',  # dk1 (text)
        2: 'e7e6e6',  # lt2
        3: '44546a',  # dk2
        4: '4472c4',  # accent1
        5: 'ed7d31',  # accent2 (orange)
        6: 'a5a5a5',  # accent3
        7: 'ffc000',  # accent4 (yellow)
        8: '5b9bd5',  # accent5 (blue)
        9: '70ad47',  # accent6 (green)
    }
    return theme_colors.get(theme_id)


def get_cell_colors(cell):
    """Extract background and text colors from cell formatting."""
    from openpyxl.styles.colors import COLOR_INDEX

    bg_color = None
    text_color = None

    # Get background color from cell fill
    if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
        color = cell.fill.fgColor or cell.fill.start_color
        if color:
            if color.type == 'indexed':
                # Handle indexed colors
                idx = color.indexed
                if idx is not None and idx < len(COLOR_INDEX):
                    rgb = COLOR_INDEX[idx]
                    if rgb and rgb != '00000000':
                        bg_color = '#' + (rgb[2:].lower() if len(rgb) == 8 else rgb.lower())
            elif color.type == 'theme':
                # Handle theme colors
                theme_rgb = get_theme_color(color.theme)
                if theme_rgb:
                    bg_color = '#' + theme_rgb
            elif color.type == 'rgb' and color.rgb:
                rgb = color.rgb
                # Excel stores ARGB (8 chars) or RGB (6 chars)
                if isinstance(rgb, str) and rgb != '00000000':
                    if len(rgb) == 8:  # ARGB format
                        bg_color = '#' + rgb[2:].lower()  # Skip alpha
                    elif len(rgb) == 6:
                        bg_color = '#' + rgb.lower()

    # Get text color from font
    if cell.font and cell.font.color:
        color = cell.font.color
        if color:
            if color.type == 'indexed':
                idx = color.indexed
                if idx is not None and idx < len(COLOR_INDEX):
                    rgb = COLOR_INDEX[idx]
                    if rgb and rgb != '00000000':
                        text_color = '#' + (rgb[2:].lower() if len(rgb) == 8 else rgb.lower())
            elif color.type == 'theme':
                theme_rgb = get_theme_color(color.theme)
                if theme_rgb:
                    text_color = '#' + theme_rgb
            elif color.type == 'rgb' and color.rgb:
                rgb = color.rgb
                if isinstance(rgb, str) and rgb != '00000000':
                    if len(rgb) == 8:
                        text_color = '#' + rgb[2:].lower()
                    elif len(rgb) == 6:
                        text_color = '#' + rgb.lower()

    return bg_color, text_color

# Load the workbook - need data_only=False to get formatting
# But we also need values, so we'll load twice
print("Loading workbook (for formatting)...")
wb_format = openpyxl.load_workbook("Crop Plan 2025 V20.xlsm", data_only=False)
ws_format = wb_format["Crop Chart"]

print("Loading workbook (for values)...")
wb_values = openpyxl.load_workbook("Crop Plan 2025 V20.xlsm", data_only=True)
ws_values = wb_values["Crop Chart"]

# Colors are on the Identifier column (column 1), not the Crop column
# The Identifier column has row colors that indicate crop type
color_col_index = 1  # Column A = Identifier has the colors
print(f"Using column {color_col_index} (Identifier) for colors")

# Get headers from row 2
headers = []
for col in range(1, ws_values.max_column + 1):
    header = ws_values.cell(row=2, column=col).value
    headers.append(header)

print(f"Found {len(headers)} columns")

# Extract crop data starting from row 3
crops = []
for row in range(3, ws_values.max_row + 1):
    identifier = ws_values.cell(row=row, column=1).value  # Column A = Identifier
    if not identifier:
        continue

    crop = {"id": generate_id(identifier)}

    for col, header in enumerate(headers, start=1):
        if header:
            value = ws_values.cell(row=row, column=col).value
            crop[header] = fmt_value(value)

    # Extract colors from the Identifier column (column 1) which has row colors
    color_cell = ws_format.cell(row=row, column=color_col_index)
    bg_color, text_color = get_cell_colors(color_cell)
    crop["bgColor"] = bg_color
    crop["textColor"] = text_color

    crops.append(crop)

print(f"Extracted {len(crops)} crops")

# Count crops with colors
crops_with_bg = sum(1 for c in crops if c.get("bgColor"))
crops_with_text = sum(1 for c in crops if c.get("textColor"))
print(f"Crops with background color: {crops_with_bg}")
print(f"Crops with text color: {crops_with_text}")

# Save to JSON
output_path = "tmp/crops_from_excel.json"
with open(output_path, "w") as f:
    json.dump({"crops": crops}, f, indent=2, cls=DateTimeEncoder)

print(f"Saved to {output_path}")

# Show sample with colors
print("\nSample crop with colors:")
sample = next((c for c in crops if c.get("bgColor")), crops[0] if crops else None)
if sample:
    print(f"  Crop: {sample.get('Crop')}")
    print(f"  bgColor: {sample.get('bgColor')}")
    print(f"  textColor: {sample.get('textColor')}")

# Show unique crops with colors
unique_crops = {}
for c in crops:
    crop_name = c.get("Crop")
    if crop_name and crop_name not in unique_crops:
        unique_crops[crop_name] = {
            "bgColor": c.get("bgColor"),
            "textColor": c.get("textColor")
        }

print(f"\nUnique crops: {len(unique_crops)}")
print("\nFirst 10 crops with colors:")
for i, (name, colors) in enumerate(list(unique_crops.items())[:10]):
    print(f"  {name}: bg={colors['bgColor']}, text={colors['textColor']}")
